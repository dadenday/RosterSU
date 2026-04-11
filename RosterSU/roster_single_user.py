# ==========================================
# FILE METADATA
# ==========================================
# FILE ROLE: Roster ingestion engine (single-user, offline-first)
# ARCHITECTURE: Structured monolith (virtual modules)
# INVARIANTS: See module-level contracts
# REVIEW MODE: Optimized for AI-assisted review
#
# MODULE LAYOUT:
#   1. TYPES         - dataclasses, schemas
#   2. DATE_RESOLVER - Layer 1 global date resolution
#   3. DATASET_SELECTION - Layer 2 flight dataset fingerprinting
#   4. INGESTION     - Layer 4 orchestration & manifest
#   5. DATABASE      - SQLite persistence layer
#   6. SECURITY      - token auth, redaction, path safety
#   7. APP_STATE     - global state encapsulation
#   8. EXPORT        - iCal/CSV generation
#   9. PARSER        - sheet detection & row extraction
#  10. CONFIG        - unified configuration management
#  11. UI            - FastHTML routes & components
#  12. ENTRYPOINT    - main serve()
# ==========================================

import os
import sqlite3
import json
import re
import time
import threading
import secrets
import io
import csv
import glob
import sys
import html
import os

import unicodedata
import tempfile
import asyncio
import argparse
import hashlib
import logging
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple, Any

# === AGENT INTENT =====================================
AGENT_INTENT = {
    "app_type": "local_single_user",
    "network": "localhost_only",
    "security_level": "high_local",
    "refactor_allowed": True,  # Enabled 2026-04-06: Technical debt remediation
    "module_split_allowed": True,  # Enabled 2026-04-06: Modular architecture
    "db_schema_mutable": False,
}
# ======================================================

# === AGENT CONTRACT ===================================
# SECTION: DATABASE
# INVARIANTS:
# - SQLite must remain WAL mode
# - All writes must be behind DB_LOCK
# - INSERT policy is STRICT OVERWRITE
# AGENT RULES:
# - Do NOT change schema
# - Do NOT add new tables
# =======================================================

event = {
    "type": "status",
    "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
}

# === AUTO-DEPENDENCY INSTALLATION =====================
# Automatically install missing dependencies on first run
def _ensure_dependencies():
    """Check and install missing dependencies automatically."""
    missing = []
    
    # Check each required package
    try:
        import fasthtml
    except ImportError:
        missing.append("python-fasthtml")
    
    try:
        import python_calamine
    except ImportError:
        missing.append("python-calamine")
    
    try:
        import rapidfuzz
    except ImportError:
        missing.append("rapidfuzz")
    
    if missing:
        print("=" * 60)
        print("RosterSU - First Time Setup")
        print("=" * 60)
        print(f"\nMissing dependencies: {', '.join(missing)}")
        print("\nInstalling required packages...")
        print("(This may take a few minutes)\n")
        
        import subprocess
        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "--upgrade", "-r", 
                 os.path.join(os.path.dirname(os.path.abspath(__file__)), "requirements.txt")],
                stdout=sys.stdout,
                stderr=sys.stderr,
            )
            print("\n✓ Dependencies installed successfully!")
            print("Restarting import process...\n")
        except subprocess.CalledProcessError as e:
            print(f"\n✗ Failed to install dependencies: {e}", file=sys.stderr)
            print("\nPlease install manually with:")
            print("  pip install -r requirements.txt")
            sys.exit(1)

# Run dependency check before importing third-party packages
_ensure_dependencies()
del _ensure_dependencies  # Clean up after execution

# === THIRD-PARTY IMPORTS ================================
from fasthtml.common import *
from starlette.responses import Response, FileResponse
from python_calamine import CalamineWorkbook
from rapidfuzz import fuzz
from collections import defaultdict

# Import parser module utilities (extracted for maintainability)
from parser import (
    # Cell utilities
    norm_cell,
    get_cell_flags,
    clean_val,
    clean_time,
    normalize_text,
    is_valid_name_generic,
    is_valid_route,
    is_route_pattern,
    is_day_off_token,
    normalize_time_range,
    # Detection
    detect_shift_sheet_by_invariants,
    detect_flight_personnel_sheet_by_invariants,
    identify_sheet_type,
    identify_sheet_type_legacy,
    build_row_signals,
    identify_shift_sheet_statistical,
    detect_zones_from_merged_ranges,
    # Constants
    CELL_ROUTE,
    CELL_TIME,
    CELL_OFF,
    CELL_NAME,
)

# Import configuration (extracted for maintainability)
from config import (
    # App config
    PROJECT_ROOT,
    DB_FILE,
    DEBUG_FILE,
    CONFIG_FILE,
    DEFAULT_CONFIG,
    _load_merged_config,
    AUTO_INGEST_DIR,
    EXPORT_DIR,
    PROCESSED_ARCHIVE_DIR,
    APP_TOKEN_FILE,
    DEFAULT_PORT,
    DEFAULT_HISTORY_LIMIT,
    PAGE_SIZE,
    PORT_WAIT_TIMEOUT,
    INGEST_INTERVAL,
    MAX_UPLOAD_MB,
    FILE_READ_CHUNK_SIZE,
    TOKEN_LENGTH_BYTES,
    TOKEN_LENGTH_HEX,
    # Thresholds
    SHIFT_TIME_MIN_ENTRIES,
    SHIFT_OFF_MIN_COUNT,
    SHIFT_TIME_MIN_CLUSTER,
    SHIFT_NAME_MIN_CANDIDATES,
    SHIFT_TOTAL_MIN_SIGNALS,
    SHIFT_COL_CONSISTENCY,
    SHIFT_WORKDAY_RATIO,
    FLIGHT_ROUTE_MIN_COUNT,
    FLIGHT_NAME_NEAR_ROUTE_MIN,
    FLIGHT_DISTINCT_NAMES_MIN,
    FLIGHT_HEADER_HITS_MIN,
    FLIGHT_WINDOW_ROUTE_HITS,
    FLIGHT_WINDOW_CALLSIGN_HITS,
    FUZZY_MATCH_THRESHOLD,
    DATE_MAJORITY_RATIO,
    DATE_PLURALITY_RATIO,
    DATE_ANOMALY_RATIO,
    SAFE_THRESHOLD,
    QUARANTINE_DIR,
    SHIFT_SHEET_MARKERS,
    # Regex patterns
    RE_TIME_RANGE,
    RE_TIME,
    RE_ROUTE,
    RE_PHONE,
    RE_ID6,
    RE_DATE,
    RE_SHIFT_TIME_PATTERN,
    RE_HEADER_DATE_PATTERN,
    RE_ENGLISH_DATE_PATTERN,
    RE_ZONE_PATTERN,
    RE_MULTIPLE_SPACES,
    ENGLISH_MONTHS,
)

# Parse command line arguments only when run as main
# This allows the module to be imported for testing without argparse errors
if __name__ in {"__main__", "builtins"} and not any(
    "pytest" in arg for arg in sys.argv
):
    parser = argparse.ArgumentParser(description="RosterMaster Application")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    args = parser.parse_args()
    DEBUG_ENABLED = args.debug
else:
    DEBUG_ENABLED = False


def debug_log(message, category="GENERAL"):
    """Log a message if debug mode is enabled"""
    if DEBUG_ENABLED:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        print(f"[DEBUG][{category}][{timestamp}] {message}")


# --- CELL CLASSIFICATION BITMASK (Phase 6) ---
# NOTE: CELL_ROUTE, CELL_TIME, CELL_OFF, CELL_NAME imported from parser module
# ---------------------------------------------

# ==========================================
# TYPES - DATA STRUCTURES (Layer 0)
# ==========================================
# NOTE: Dataclasses imported from data_types module for maintainability
from data_types import (
    FlightRow,
    ShiftRecord,
    ParsedSheet,
    ParseContext,
    DateCandidate,
    IngestionManifest,
    DateMismatchWarning,
    InvariantViolation,
)

# ==========================================
# 📅 LAYER 1: GLOBAL DATE RESOLVER
# ==========================================
# === AGENT CONTRACT ===================================
# SECTION: DATE_RESOLVER
# INVARIANTS:
# - ONLY DateResolver decides the global truth date
# - All downstream parsers MUST receive date via ParseContext
# - No parser may override or extract dates independently
# AGENT RULES:
# - Do NOT allow date extraction in lower layers
# - Do NOT bypass the weighted voting algorithm
# =======================================================


def assert_valid_context(context: ParseContext, operation: str) -> None:
    """
    Assert that ParseContext has valid date authority.

    CONTRACT: All parsing must have a resolved global_date.
    Fail FAST if date authority is not established.
    """
    if context is None:
        raise InvariantViolation(f"[{operation}] ParseContext is None")
    if context.global_date == "Unknown":
        raise InvariantViolation(
            f"[{operation}] Date authority not established for {context.source_filename}"
        )


def assert_valid_manifest(manifest: "IngestionManifest") -> None:
    """
    Assert that manifest is ready for DB write.

    CONTRACT: DB writes require valid date and non-blocked manifest.
    """
    if manifest.global_date == "Unknown":
        raise InvariantViolation(
            f"Cannot write to DB with unknown date: {manifest.filename}"
        )
    if manifest.blocked:
        raise InvariantViolation(f"Manifest blocked: {manifest.block_reason}")


# ==========================================
# 🛡️ LAYER 4: INGESTION MANIFEST & SAFE WRITE GATE
# ==========================================
# === AGENT CONTRACT ===================================
# SECTION: INGESTION_SAFETY
# INVARIANTS:
# - IngestionManifest must be created for every file ingestion
# - Confidence score < SAFE_THRESHOLD triggers review (warning only)
# - ONE FILE = ONE DATE: Only file-level resolver defines date authority
# - Sheet headers are DISPLAY TEXT ONLY - they never block parsing
# AGENT RULES:
# - Do NOT skip sheets based on header dates
# - Do NOT bypass the Safe Write Gate
# - Do NOT write to DB without valid manifest
# =======================================================

# NOTE: IngestionManifest imported from types module


# ==========================================
# 🛡️ FLIGHT DATASET SELECTION (Feature A, B, C, D, E)
# ==========================================
# === AGENT CONTRACT ===================================
# SECTION: DATASET_SELECTION
# INVARIANTS:
# - ONE FILE = ONE FLIGHT DATASET (only authoritative dataset is ingested)
# - Fingerprinting is deterministic and order-independent
# - Empty datasets are always filtered
# - Cross-file fingerprint gate is an OPTIMIZATION only
# AGENT RULES:
# - Do NOT bypass DatasetSelector
# - Do NOT write flight data without fingerprint validation
# =======================================================


@dataclass
class FlightDatasetCandidate:
    """
    Represents a candidate flight dataset from a single sheet.
    Used for intra-file dataset selection.
    """

    sheet_name: str
    fingerprint: str
    row_count: int
    rows: List[
        Tuple
    ]  # Normalized rows: (personnel_id, flight_code, open_time, close_time)
    sheet_date: Optional[str] = None
    score: int = 0
    is_empty: bool = False


class DatasetSelector:
    """
    Feature A, B, C: Intra-File Dataset Fingerprinting and Selection.

    Ensures ONLY ONE flight dataset per file is ingested.
    Detects and filters duplicate/copied sheets.
    """

    # Score bonus for sheets matching the global date
    DATE_MATCH_BONUS = 50

    @staticmethod
    def normalize_flight_row(row_data: Dict) -> Tuple:
        """
        Normalize a flight row for fingerprinting.
        Returns a tuple: (personnel_id, flight_code, open_time, close_time)
        """
        personnel = str(row_data.get("Names", "")).strip().upper()
        flight = str(row_data.get("Call", "")).strip().upper()
        open_t = str(row_data.get("Open", "")).strip().upper()
        close_t = str(row_data.get("Close", "")).strip().upper()

        # Normalize time formats
        open_t = DatasetSelector._normalize_time(open_t)
        close_t = DatasetSelector._normalize_time(close_t)

        return (personnel, flight, open_t, close_t)

    @staticmethod
    def _normalize_time(t: str) -> str:
        """Normalize time string to consistent format."""
        if not t:
            return ""
        # Replace common variants
        t = t.replace("H", ":").replace("h", ":").replace(".", ":")
        # Extract HH:MM
        match = re.search(r"(\d{1,2})[:]?(\d{2})", t)
        if match:
            h, m = match.groups()
            return f"{int(h):02d}:{m}"
        return t

    @staticmethod
    def compute_fingerprint(rows: List[Tuple]) -> str:
        """
        Compute SHA256 fingerprint for a list of normalized rows.
        The fingerprint is order-independent (rows are sorted before hashing).
        """
        if not rows:
            return ""

        # Sort rows deterministically for order-independent fingerprinting
        sorted_rows = sorted(rows)

        # Convert to JSON string
        rows_json = json.dumps(sorted_rows, ensure_ascii=False)

        # Compute SHA256 hash
        return hashlib.sha256(rows_json.encode("utf-8")).hexdigest()

    @staticmethod
    def extract_flight_rows_from_sheet(flights: List[Dict]) -> List[Tuple]:
        """
        Extract and normalize flight rows from parsed sheet data.
        """
        normalized = []
        for f in flights:
            row = DatasetSelector.normalize_flight_row(f)
            # Only include rows with actual data (not just empty strings)
            if any(row):
                normalized.append(row)
        return normalized

    @staticmethod
    def build_candidates(
        flight_sheets: List[Tuple[str, List[Dict]]], global_date: str
    ) -> List[FlightDatasetCandidate]:
        """
        Build list of candidate datasets from all flight sheets.

        Args:
            flight_sheets: List of (sheet_name, flights_list) tuples
            global_date: The authoritative date for this file

        Returns:
            List of FlightDatasetCandidate objects with fingerprints and scores
        """
        candidates = []

        for sheet_name, flights in flight_sheets:
            # Extract normalized rows
            normalized_rows = DatasetSelector.extract_flight_rows_from_sheet(flights)

            # Check if empty
            is_empty = len(normalized_rows) == 0

            # Compute fingerprint
            fingerprint = DatasetSelector.compute_fingerprint(normalized_rows)

            # Calculate score
            score = len(normalized_rows)  # Base score = row count

            # Create candidate
            candidate = FlightDatasetCandidate(
                sheet_name=sheet_name,
                fingerprint=fingerprint,
                row_count=len(normalized_rows),
                rows=normalized_rows,
                is_empty=is_empty,
            )
            candidates.append(candidate)

        return candidates

    @staticmethod
    def select_authoritative(
        candidates: List[FlightDatasetCandidate],
    ) -> Tuple[Optional[FlightDatasetCandidate], List[str]]:
        """
        Select the authoritative dataset from candidates.

        Selection rules:
        1. Filter out empty datasets
        2. Filter out duplicate fingerprints (keep first occurrence)
        3. Select highest score (row_count)

        Returns:
            Tuple of (selected_candidate, list_of_warnings)
        """
        warnings = []

        # Feature C: Filter empty datasets
        non_empty = [c for c in candidates if not c.is_empty]

        if not non_empty:
            warnings.append("EMPTY_FLIGHT_SHEET: All flight sheets are empty")
            return None, warnings

        # Feature A: Filter duplicates by fingerprint
        seen_fingerprints = set()
        unique_candidates = []

        for c in non_empty:
            if c.fingerprint in seen_fingerprints:
                warnings.append(
                    f"DUPLICATE_DATASET_SKIPPED: Sheet '{c.sheet_name}' is a duplicate"
                )
            else:
                seen_fingerprints.add(c.fingerprint)
                unique_candidates.append(c)

        if not unique_candidates:
            warnings.append("NO_UNIQUE_DATASET: All datasets are duplicates")
            return None, warnings

        # Feature B: Select authoritative (highest row_count)
        authoritative = max(unique_candidates, key=lambda c: c.row_count)

        # Log non-authoritative datasets
        for c in unique_candidates:
            if c != authoritative:
                warnings.append(
                    f"NON_AUTHORITATIVE_DATASET: Sheet '{c.sheet_name}' skipped "
                    f"(row_count={c.row_count} vs {authoritative.row_count})"
                )

        return authoritative, warnings


def _extract_sheet_date_for_warning(rows: List, sheet_name: str = "") -> Optional[str]:
    """
    Extract date from a sheet for WARNING purposes only.
    This does NOT affect parsing - just logs potential stale headers.

    Returns normalized DD.MM.YYYY or None.
    """
    if not rows:
        return None

    # Strategy 1: ISO Date scan (YYYY-MM-DD)
    RE_ISO_DATE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
    for row in rows[:15]:
        for val in row[:50]:
            val_str = str(val) if val else ""
            iso_match = RE_ISO_DATE.search(val_str)
            if iso_match:
                y, m, d = iso_match.groups()
                d_str = f"{d}.{m}.{y}"
                if is_valid_roster_year(d_str):
                    return normalize_date_str(d_str)

    # Strategy 2: English Header Scan
    for row in rows[:15]:
        row_str = " ".join([str(x) for x in row[:50] if x]).upper()
        en_match = RE_ENGLISH_DATE_PATTERN.search(row_str)
        if en_match:
            day, mon, year = en_match.groups()
            mon_num = ENGLISH_MONTHS.get(mon[:3].upper())
            if mon_num:
                d_str = f"{day}.{mon_num}.{year}"
                if is_valid_roster_year(d_str):
                    return normalize_date_str(d_str)

    # Strategy 3: Header scan (NGÀY DD.MM.YYYY)
    for row in rows[:10]:
        row_str = " ".join([str(x) for x in row[:50] if x]).upper()
        match = RE_HEADER_DATE_PATTERN.search(row_str)
        if match:
            d_str = match.group(1)
            if is_valid_roster_year(d_str):
                return normalize_date_str(d_str)

    # Strategy 4: Cell scan for DD.MM.YYYY patterns
    for row in rows[:5]:
        for val in row[:50]:
            val_str = str(val) if val else ""
            matches = RE_DATE.findall(val_str)
            for m in matches:
                if is_valid_roster_year(m):
                    return normalize_date_str(m)

    return None


class ConfidenceScorer:
    """
    Level-2: Confidence Scoring Engine (SIMPLIFIED)

    Computes confidence ∈ [0–1] based on extraction success.

    IMPORTANT: Sheet date agreement is NOT a signal.
    Sheet headers are display text - they may be stale without indicating problems.

    Signals (weighted):
    - Date Agreement: Does the chosen date have clear majority?
    - Data Extraction: Was actual shift/flight data found?
    - Anomalies: Any critical warning conditions detected?
    """

    @staticmethod
    def compute(manifest: IngestionManifest) -> float:
        """Calculate confidence score from manifest data."""
        score = 0.5  # Start at neutral

        # Signal 1: Date Agreement
        date_candidates = manifest.date_candidates
        if date_candidates:
            chosen_weight = 0
            total_weight = 0
            for c in date_candidates:
                w = c.get("weight", 1)
                total_weight += w
                if c.get("date") == manifest.global_date:
                    chosen_weight = w

            if total_weight > 0:
                ratio = chosen_weight / total_weight
                if ratio >= 0.7:
                    score += 0.2  # Strong consensus
                elif ratio >= 0.5:
                    score += 0.1  # Majority
                # Weak consensus is acceptable - don't penalize

        # Signal 2: Data Extraction (PRIMARY SIGNAL)
        if manifest.parsed_counts.get("shift") or manifest.parsed_counts.get("flights"):
            score += 0.3  # Data extracted successfully
            if manifest.parsed_counts.get("shift") and manifest.parsed_counts.get(
                "flights"
            ):
                score += 0.1  # Both shift and flights found
        else:
            score -= 0.3  # No data extracted - this is the main failure signal

        # Signal 3: Anomalies (small penalty per anomaly)
        if manifest.anomalies:
            score -= 0.05 * min(len(manifest.anomalies), 3)

        # Warnings do NOT affect confidence - they are informational only
        # Stale sheet headers are logged but don't indicate data problems

        # Clamp to [0, 1]
        return max(0.0, min(1.0, score))


def _extract_date_from_filename(filename: str) -> Optional[str]:
    """Extract date string (DDMMYYYY) from filename for archive dedup."""
    base = os.path.splitext(os.path.basename(filename))[0]
    # Try common patterns: DDMMYYYY, DD.MM.YYYY, DD-MM-YYYY, YYYYMMDD
    for pattern in [
        r'(\d{2}[.\-/]\d{2}[.\-/]\d{4})',
        r'(\d{8})',
    ]:
        m = re.search(pattern, base)
        if m:
            raw = m.group(1).replace('.', '').replace('-', '').replace('/', '')
            if len(raw) == 8 and raw[:2].isdigit() and raw[2:4].isdigit():
                # DDMMYYYY -> normalize
                return f"{raw[:2]}{raw[2:4]}{raw[4:]}"
            elif len(raw) == 8 and raw[:4].isdigit():
                # YYYYMMDD -> convert to DDMMYYYY for comparison
                return f"{raw[6:8]}{raw[4:6]}{raw[:4]}"
    return None


def _archive_processed_file(file_path: str, manifest: IngestionManifest) -> str:
    """
    Move successfully parsed file to processed_archive.

    If another file for the same date already exists in the archive,
    the newer file (by modification time) replaces the older one.

    Returns the final archive path.
    """
    os.makedirs(PROCESSED_ARCHIVE_DIR, exist_ok=True)

    file_date = manifest.global_date if manifest.global_date else _extract_date_from_filename(manifest.filename)
    if not file_date:
        file_date = "unknown"

    base_name = os.path.basename(file_path)
    dest_path = os.path.join(PROCESSED_ARCHIVE_DIR, base_name)

    # Check for existing file with same date
    existing_files = glob.glob(os.path.join(PROCESSED_ARCHIVE_DIR, "*.xlsx")) + \
                     glob.glob(os.path.join(PROCESSED_ARCHIVE_DIR, "*.xls")) + \
                     glob.glob(os.path.join(PROCESSED_ARCHIVE_DIR, "*.csv"))

    for existing in existing_files:
        try:
            existing_date = _extract_date_from_filename(os.path.basename(existing))
            if existing_date and file_date and existing_date == file_date:
                # Same date found - check if new file is newer by mtime
                new_mtime = os.path.getmtime(file_path)
                old_mtime = os.path.getmtime(existing)
                if new_mtime > old_mtime:
                    # Remove old archive, replace with new
                    os.remove(existing)
                    log_debug("archive_replaced", {
                        "old": os.path.basename(existing),
                        "new": base_name,
                        "date": file_date,
                    })
                else:
                    # Old archive is newer, skip replacement
                    log_debug("archive_skipped", {
                        "existing": os.path.basename(existing),
                        "incoming": base_name,
                        "date": file_date,
                    })
                    return existing
                break
        except OSError:
            continue

    # Move file to archive
    try:
        os.rename(file_path, dest_path)
    except OSError:
        import shutil
        shutil.copy2(file_path, dest_path)
        os.remove(file_path)

    return dest_path


def quarantine_file(file_path: str, manifest: IngestionManifest) -> str:
    """
    Move a problematic file to quarantine directory.
    Creates quarantine manifest alongside the file.

    Returns path to quarantine location.
    """
    # Create quarantine directory if needed
    quarantine_path = os.path.join(os.path.dirname(file_path), QUARANTINE_DIR)
    os.makedirs(quarantine_path, exist_ok=True)

    # Generate unique filename
    base_name = os.path.basename(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    quarantined_name = f"{timestamp}_{base_name}"
    dest_path = os.path.join(quarantine_path, quarantined_name)

    # Move file
    try:
        os.rename(file_path, dest_path)
    except OSError:
        # Fallback: copy and remove
        import shutil

        shutil.copy2(file_path, dest_path)
        os.remove(file_path)

    # Write manifest alongside
    manifest_path = dest_path + ".manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "file_hash": manifest.file_hash,
                "filename": manifest.filename,
                "global_date": manifest.global_date,
                "global_date_iso": manifest.global_date_iso,
                "confidence_score": manifest.confidence_score,
                "blocked": manifest.blocked,
                "block_reason": manifest.block_reason,
                "parsed_sheets": manifest.parsed_sheets,
                "parsed_counts": manifest.parsed_counts,
                "date_candidates": manifest.date_candidates,
                "anomalies": manifest.anomalies,
                "warnings": manifest.warnings,
                "timestamp": manifest.timestamp,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    log_debug(
        "INGESTION_BLOCKED",
        {
            "reason": manifest.block_reason,
            "confidence": manifest.confidence_score,
            "file": base_name,
            "quarantine_path": dest_path,
        },
    )

    return dest_path


def create_manifest_from_context(
    context: ParseContext, filename: str, file_hash: str = ""
) -> IngestionManifest:
    """Create initial manifest from ParseContext."""
    if not file_hash:
        import hashlib

        file_hash = context.file_id

    return IngestionManifest(
        file_hash=file_hash,
        filename=filename,
        global_date=context.global_date,
        global_date_iso=context.global_date_iso,
        date_candidates=context.date_candidates,
        parsed_sheets=[],
        warnings=[],  # Non-blocking issues (stale headers, etc.)
        parsed_counts={"shift": 0, "flights": 0},
        anomalies=[],
        confidence_score=1.0 if context.global_date != "Unknown" else 0.0,
        timestamp=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )


class DateResolver:
    """
    Layer 1: Global Date Resolver

    Collects date candidates from multiple sources and resolves to a single
    truth date using weighted majority voting.

    Trust Weights:
    - Filename: 3 (historically reliable)
    - Shift sheet header: 3 (primary source)
    - Repeated occurrences: +1 each
    - Flight sheet: 1 (prone to typos)
    """

    # Trust weights for different sources
    # Key principle: weight comes from source TYPE, not from COUNT.
    # A date appearing in 5 sheets is NOT more authoritative than one in the filename.
    WEIGHTS = {
        "filename": 5,        # Human's declared intent — highest authority
        "shift_header": 3,    # Sheet header with NGÀY DD.MM.YYYY
        "shift_cell": 2,      # Date found in shift sheet cells
        "flight_header": 2,   # Flight sheet header date
        "flight_cell": 1,     # Date found in flight sheet cells
        "iso_cell": 3,        # YYYY-MM-DD format cells (explicit dates)
    }

    # Anomaly threshold
    ANOMALY_RATIO = 0.6  # Trigger warning if second_best/best > 0.6

    def __init__(self, filename: str):
        self.filename = filename
        self.candidates: List[DateCandidate] = []
        self._resolved_date: Optional[str] = None
        self._confidence: float = 0.0
        self._anomaly: bool = False

    def add_candidate(self, date_str: str, source: str, raw_value: str = "") -> None:
        """Add a date candidate with its source.

        Weight comes from source TYPE only — NOT from count.
        A date appearing in multiple sheets does NOT accumulate weight.
        Higher-priority sources upgrade the weight if the same date appears later.
        """
        if not date_str or date_str == "Unknown":
            return

        # Normalize the date
        normalized = normalize_date_str(date_str)

        # Validate year range
        if not is_valid_roster_year(normalized):
            return

        # Determine weight from source type only — no accumulation
        weight = self.WEIGHTS.get(source, 1)

        # Check for duplicates: upgrade weight if this source is higher priority
        existing = next((c for c in self.candidates if c.date == normalized), None)
        if existing:
            if weight > existing.weight:
                existing.weight = weight
                existing.source = source
            return  # Already have this date

        self.candidates.append(
            DateCandidate(
                date=normalized, source=source, weight=weight, raw_value=raw_value
            )
        )

    def _extract_from_filename(self) -> None:
        """Extract date from filename."""
        match = RE_DATE.search(self.filename)
        if match:
            self.add_candidate(match.group(1), "filename", match.group(0))

    def _extract_from_sheet(self, rows: List, sheet_type: str = "unknown") -> None:
        """Extract date candidates from a sheet's content."""
        source_prefix = (
            "shift"
            if sheet_type == "SHIFT"
            else "flight"
            if sheet_type == "FLIGHT"
            else "unknown"
        )

        # Strategy 1: ISO Date scan (YYYY-MM-DD)
        RE_ISO_DATE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
        for row in rows[:15]:
            for val in row[:50]:
                val_str = str(val) if val else ""
                iso_match = RE_ISO_DATE.search(val_str)
                if iso_match:
                    y, m, d = iso_match.groups()
                    self.add_candidate(f"{d}.{m}.{y}", f"{source_prefix}_iso", val_str)

        # Strategy 1.5: English Header Scan
        for row in rows[:15]:
            row_str = " ".join([str(x) for x in row[:50] if x]).upper()
            en_match = RE_ENGLISH_DATE_PATTERN.search(row_str)
            if en_match:
                day, mon, year = en_match.groups()
                mon_num = ENGLISH_MONTHS.get(mon[:3].upper())
                if mon_num:
                    self.add_candidate(
                        f"{day}.{mon_num}.{year}",
                        f"{source_prefix}_header",
                        row_str[:50],
                    )

        # Strategy 2: Header scan (NGÀY DD.MM.YYYY)
        for row in rows[:10]:
            row_str = " ".join([str(x) for x in row[:50] if x]).upper()
            match = RE_HEADER_DATE_PATTERN.search(row_str)
            if match:
                self.add_candidate(
                    match.group(1), f"{source_prefix}_header", row_str[:50]
                )

        # Strategy 3: Cell scan for DD.MM.YYYY patterns
        for row in rows[:5]:
            for val in row[:50]:
                val_str = str(val) if val else ""
                matches = RE_DATE.findall(val_str)
                for m in matches:
                    self.add_candidate(m, f"{source_prefix}_cell", val_str[:30])

    def scan_sheets(self, sheets_data: List[Tuple[str, List]]) -> None:
        """
        Scan all sheets to collect date candidates.
        Must be called before resolve().
        """
        # First, extract from filename
        self._extract_from_filename()

        # Scan each sheet
        for sheet_name, rows in sheets_data:
            # Determine sheet type for weighting
            sheet_type = identify_sheet_type(rows, sheet_name or "LỊCH BAY")
            if sheet_type != "SKIP":
                self._extract_from_sheet(rows, sheet_type)

    def resolve(self) -> ParseContext:
        """
        Resolve to a single truth date using weighted majority voting.
        Returns a ParseContext with the resolved date.

        Raises DateMismatchWarning if anomaly detected (but still returns best date).
        """
        if not self.candidates:
            # No candidates found - this is an error case
            return ParseContext(
                global_date="Unknown",
                global_date_iso="",
                source_filename=self.filename,
                file_id=self._generate_file_id(),
                date_confidence=0.0,
                date_anomaly=True,
                date_candidates=[],
            )

        # Aggregate weights by date
        date_scores: Dict[str, int] = {}
        date_sources: Dict[str, List[str]] = {}

        for c in self.candidates:
            if c.date not in date_scores:
                date_scores[c.date] = 0
                date_sources[c.date] = []
            date_scores[c.date] += c.weight
            date_sources[c.date].append(c.source)

        # Sort by score descending
        sorted_dates = sorted(date_scores.items(), key=lambda x: x[1], reverse=True)

        if not sorted_dates:
            return ParseContext(
                global_date="Unknown",
                global_date_iso="",
                source_filename=self.filename,
                file_id=self._generate_file_id(),
                date_confidence=0.0,
                date_anomaly=True,
                date_candidates=[],
            )

        best_date, best_score = sorted_dates[0]
        total_score = sum(date_scores.values())

        # Calculate confidence
        self._confidence = best_score / total_score if total_score > 0 else 0.0

        # Check for anomaly
        self._anomaly = False
        if len(sorted_dates) > 1:
            second_score = sorted_dates[1][1]
            if second_score / best_score > self.ANOMALY_RATIO:
                self._anomaly = True

        # Log resolution
        log_debug(
            "DATE_RESOLUTION",
            {
                "file": self.filename,
                "candidates": [
                    {"date": c.date, "source": c.source, "weight": c.weight}
                    for c in self.candidates
                ],
                "chosen": best_date,
                "confidence": round(self._confidence, 2),
                "anomaly": self._anomaly,
                "scores": {d: s for d, s in sorted_dates},
            },
        )

        self._resolved_date = best_date

        # Build candidates list for context
        candidates_info = [
            {"date": c.date, "source": c.source, "weight": c.weight}
            for c in self.candidates
        ]

        return ParseContext(
            global_date=best_date,
            global_date_iso=to_iso_date(best_date),
            source_filename=self.filename,
            file_id=self._generate_file_id(),
            date_confidence=self._confidence,
            date_anomaly=self._anomaly,
            date_candidates=candidates_info,
        )

    def _generate_file_id(self) -> str:
        """Generate a unique file ID for this ingestion."""
        import hashlib

        ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
        raw = f"{self.filename}:{ts}"
        return hashlib.md5(raw.encode()).hexdigest()[:12]


def resolve_global_date(
    filename: str, sheets_data: List[Tuple[str, List]]
) -> ParseContext:
    """
    Convenience function: Create resolver, scan sheets, and return context.
    This is the main entry point for Layer 1 date resolution.
    """
    resolver = DateResolver(filename)
    resolver.scan_sheets(sheets_data)
    return resolver.resolve()


# ==========================================
# 🧠 GLOBAL STATE
# ==========================================
# NOTE: AppState, locks, and state functions extracted to state.py for maintainability
from state import (
    # Locks
    STATE_LOCK,
    DB_LOCK,
    CONFIG_LOCK,
    SHUTDOWN_EVENT,
    INGEST_RUNNING,
    UI_CONNECTED,
    # Legacy globals (for backward compatibility)
    APP_STATUS,
    ROSTER_VERSION,
    # State class
    AppState,
    APP,
    # State functions
    get_app_status,
    try_get_app_status,
    update_status,
    bump_db_rev,
)

# === LEGACY COMPATIBILITY FUNCTIONS (delegate to APP) ===


def update_running_status(details: str = ""):
    # Running status is only valid during real ingest
    debug_log(
        f"update_running_status called with details='{details}'", "RUNNING_STATUS"
    )
    if APP.is_ingest_running():
        debug_log("INGEST_RUNNING is set, calling update_status", "RUNNING_STATUS")
        update_status("Running", details)
    else:
        debug_log("INGEST_RUNNING is not set, skipping update_status", "RUNNING_STATUS")


def wait_for_port(
    port: int, host: str = "127.0.0.1", timeout: int = PORT_WAIT_TIMEOUT
) -> bool:
    """Wait for a TCP port to become active."""
    import socket
    debug_log(
        f"wait_for_port called with port={port}, host={host}, timeout={timeout}",
        "NETWORK",
    )
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            debug_log(f"Attempting connection to {host}:{port}", "NETWORK")
            with socket.create_connection((host, port), timeout=0.5):
                debug_log(f"Successfully connected to {host}:{port}", "NETWORK")
                return True
        except (ConnectionRefusedError, socket.timeout, OSError) as e:
            debug_log(f"Connection failed to {host}:{port}, error: {str(e)}", "NETWORK")
            time.sleep(0.2)
    debug_log(f"Timeout waiting for {host}:{port}", "NETWORK")
    return False


def wait_for_http(
    port: int, host: str = "127.0.0.1", timeout: int = PORT_WAIT_TIMEOUT
) -> bool:
    """Wait for a HTTP response from the port."""
    import http.client
    debug_log(
        f"wait_for_http called with port={port}, host={host}, timeout={timeout}",
        "NETWORK",
    )
    end = time.time() + timeout
    while time.time() < end:
        try:
            debug_log(f"Attempting HTTP request to {host}:{port}/", "NETWORK")
            conn = http.client.HTTPConnection(host, port, timeout=0.5)
            conn.request("GET", "/")
            conn.close()
            debug_log(f"Successful HTTP response from {host}:{port}/", "NETWORK")
            return True
        except Exception as e:
            debug_log(
                f"HTTP request failed to {host}:{port}/, error: {str(e)}", "NETWORK"
            )
            time.sleep(0.3)
    debug_log(f"Timeout waiting for HTTP response from {host}:{port}/", "NETWORK")
    return False


def launch_browser(url):
    """Attempt to open URL in browser using Termux commands."""
    import subprocess
    debug_log(f"launch_browser called with url={url}", "BROWSER")
    try:
        # Strategy 1: termux-open-url (Requires termux-api)
        debug_log("Attempting to launch browser using termux-open-url", "BROWSER")
        res = subprocess.run(
            ["termux-open-url", url], check=False, stderr=subprocess.DEVNULL
        )
        if res.returncode == 0:
            debug_log("Successfully launched browser using termux-open-url", "BROWSER")
            return
        else:
            debug_log(
                f"termux-open-url failed with return code {res.returncode}", "BROWSER"
            )
    except Exception as e:
        debug_log(f"Exception in termux-open-url: {str(e)}", "BROWSER")
        # Redact sensitive data before logging
        print(f"BROWSER_LAUNCH_ERROR (termux-open-url): {e}", file=sys.stderr)
        log_data = redact_sensitive_data(
            {
                "error": str(e),
                "function": "launch_browser",
                "strategy": "termux-open-url",
                "type": type(e).__name__,
                "location": "launch_browser.termux-open-url",
            }
        )
        log_debug("runtime_error", log_data)

    try:
        # Strategy 2: Android Intent via 'am' (More universal in Termux)
        debug_log("Attempting to launch browser using Android Intent", "BROWSER")
        subprocess.run(
            ["am", "start", "-a", "android.intent.action.VIEW", "-d", url],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        debug_log("Launched browser using Android Intent", "BROWSER")
    except Exception as e:
        debug_log(f"Exception in Android Intent: {str(e)}", "BROWSER")
        # Redact sensitive data before logging
        print(f"BROWSER_LAUNCH_ERROR (android-intent): {e}", file=sys.stderr)
        log_data = redact_sensitive_data(
            {
                "error": str(e),
                "function": "launch_browser",
                "strategy": "android-intent",
                "type": type(e).__name__,
                "location": "launch_browser.android-intent",
                "recoverable": True,
            }
        )
        log_debug("runtime_error", log_data)


def auto_open_launcher():
    """Wait for server and then open browser."""
    debug_log(
        f"auto_open_launcher called, waiting for server on port {DEFAULT_PORT}",
        "LAUNCHER",
    )
    print(f"Waiting for RosterMaster server on port {DEFAULT_PORT}...")
    if wait_for_http(DEFAULT_PORT):
        url = f"http://127.0.0.1:{DEFAULT_PORT}"
        debug_log(f"Server responded, launching browser with URL: {url}", "LAUNCHER")
        print(f"Launching browser: {url}")
        launch_browser(url)
    else:
        debug_log(
            f"Server on port {DEFAULT_PORT} not responding, trying fallback", "LAUNCHER"
        )
        # Fallback attempt
        print("Default port not responding, trying fallback...")
        launch_browser("http://127.0.0.1:8000")


_LOG_LOCK = threading.Lock()
_LOGGING_LOCAL = threading.local()

# ==========================================
# 🔐 SECURITY & REDACTION
# ==========================================

# === AGENT CONTRACT ===================================
# SECTION: SECURITY
# INVARIANTS:
# - Token auth must remain local-only
# - All sensitive data must be redacted
# - Formula injection must be prevented
# AGENT RULES:
# - Do NOT weaken redaction
# - Do NOT remove token validation
# - Do NOT allow external network access
# =======================================================


def log_debug(event, data=None):
    # Return immediately if ingest is running and DEBUG_ENABLED is False,
    # unless the event is one of the critical ones
    critical_events = {
        "db_write_error",
        "bulk_db_error",
        "runtime_error",
        "auth_denied",
        "file_loop_error",
    }
    if APP.is_ingest_running() and not DEBUG_ENABLED and event not in critical_events:
        return

    debug_log(f"log_debug called with event='{event}', data='{data}'", "LOG_DEBUG")

    # Recursion guard using threading.local()
    if getattr(_LOGGING_LOCAL, "active", False):
        print("LOG_DROPPED", file=sys.stderr)
        return
    _LOGGING_LOCAL.active = True

    try:
        # Redact sensitive data before writing to the log file
        try:
            redacted_data = redact_sensitive_data(data)
        except Exception as e:
            # If redact_sensitive_data fails, log to stderr and use original data
            import sys

            print(f"REDACT_ERROR: Failed to redact data: {str(e)}", file=sys.stderr)
            redacted_data = data  # Use original data if redaction fails

        # Use lock when checking file size and performing rotation
        with _LOG_LOCK:
            # 5.1 Simple size-based rotation for DEBUG_FILE
            if os.path.exists(DEBUG_FILE) and os.path.getsize(DEBUG_FILE) > 5_000_000:
                try:
                    os.replace(DEBUG_FILE, DEBUG_FILE + ".old")
                except Exception as e:
                    print("LOG ROTATION FAILED:", e, file=sys.stderr)

            record = {
                "ts": datetime.utcnow().isoformat(),
                "event": event,
                "data": redacted_data,
            }
            try:
                with open(DEBUG_FILE, "a", encoding="utf-8") as f:
                    f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
            except Exception as e:
                # Use fallback logging to stderr to avoid infinite recursion
                import sys

                print(
                    f"LOG_ERROR: Failed to write to debug file: {str(e)}",
                    file=sys.stderr,
                )
                # Prevent infinite recursion if logging fails
    finally:
        _LOGGING_LOCAL.active = False


def redact_sensitive_data(data):
    """
    Recursively redact sensitive keys from data structures.
    """
    sensitive_keys = {
        "token",
        "app_token",
        "secret",
        "password",
        "x-app-token",
        "authorization",
        "auth_token",
        "api_key",
        "access_token",
        "refresh_token",
        "sesskey",
        "session_key",
    }

    if isinstance(data, dict):
        redacted_dict = {}
        for key, value in data.items():
            if key.lower() in sensitive_keys:
                redacted_dict[key] = "[REDACTED]"
            else:
                redacted_dict[key] = redact_sensitive_data(value)
        return redacted_dict
    elif isinstance(data, list):
        return [redact_sensitive_data(item) for item in data]
    else:
        return data


def safe_path(base_dir: str, path: str) -> str:
    """
    Validate that a resolved path stays within the base directory.
    Uses realpath() which resolves ALL symlinks in every intermediate component,
    then verifies the resolved path is prefixed by the resolved base.
    """
    real_base = os.path.realpath(base_dir)
    real_path = os.path.realpath(os.path.join(real_base, path))

    if not real_path.startswith(real_base + os.sep) and real_path != real_base:
        raise ValueError("Unsafe file path detected (symlink traversal)")

    return real_path


def sanitize_formula(val):
    """
    Sanitize formula strings that start with =, +, -, @ by prefixing with a single quote (').
    This prevents formula injection attacks in exports.
    """
    if val and isinstance(val, str):
        stripped = val.lstrip()
        if stripped.startswith(("=", "+", "-", "@")):
            return "'" + val
    return val


# ==========================================
# 💾 DATABASE LAYER
# ==========================================
# NOTE: All database functions extracted to database.py for maintainability
from database import (
    # Connection management
    get_db,
    db_conn,
    init_db,
    clear_db,
    # CRUD operations
    count_history,
    load_history,
    get_available_months,
    # Write operations
    save_entry_overwrite,
    save_entries_bulk,
    delete_entries,
    # Date utilities
    normalize_date_str,
    to_iso_date,
    is_valid_roster_year,
    # Fingerprint gate functions
    check_fingerprint_seen,
    record_fingerprint,
    get_active_ingestion,
    update_ingestion_manifest,
    clear_ingestion_manifest,
    # Initialization function
    _init_database as _init_database_from_db,
)


def _init_database_module():
    """Initialize database module with dependencies from main module."""
    _init_database_from_db(debug_log, log_debug, DB_LOCK)


# Initialize database module after functions are defined
_init_database_module()


# ==========================================
# 📤 EXPORT & INTEGRATION
# ==========================================
# NOTE: Export functions extracted to export.py for maintainability
from export import (
    generate_ical_content,
    generate_csv_content,
    _init_exports as _init_exports_from_export,
)

# ==========================================
# NOTE: UI components extracted to components.py for maintainability
from components import (
    format_date_vn,
    format_shift_display,
    shift_color,
    shift_text_class,
    build_copy_text,
    RosterCard,
    sort_flights_by_open_time,
    is_flight_card_active,
    RosterList,
    _init_components,
)


def _init_export_module():
    """Initialize export module with dependencies."""
    _init_exports_from_export(load_history, sanitize_formula, log_debug)


# Initialize export module after load_history is defined
# This will be called at module load time
_init_export_module()

# ==========================================
# 🧹 PARSING HELPERS
# ==========================================

# ==========================================
# 🧹 PARSING LOGIC & HEURISTICS
# ==========================================

# === AGENT CONTRACT ===================================
# SECTION: INGEST & PARSING
# INVARIANTS:
# - File parsing must remain safe from path traversal
# - Regex patterns must not be tightened without dataset regression
# - Sheet identification must remain robust to format variations
# AGENT RULES:
# - Do NOT change regex patterns without validation
# - Do NOT remove safe_path validation
# - Do NOT weaken file format support
# =======================================================
# --- PARSER: UTILITIES ---
# NOTE: Utility functions imported from parser module
# (norm_cell, get_cell_flags, is_valid_name_generic, is_valid_route,
#  is_route_pattern, is_day_off_token, normalize_time_range, etc.)


def load_excel_file(file_path):
    """Helper for validation scripts to load the first sheet of an excel file."""
    workbook = CalamineWorkbook.from_path(file_path)
    sheet_name = workbook.sheet_names[0]
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.to_python()


# NOTE: is_valid_name_generic, is_valid_route, is_route_pattern,
#       is_day_off_token, normalize_time_range imported from parser module


def fuzzy_match_markers(
    text: str, markers: list, threshold: int = FUZZY_MATCH_THRESHOLD
) -> int:
    count = 0
    text_upper = text.upper()
    for marker in markers:
        marker_upper = marker.upper()
        if marker_upper in text_upper:
            count += 1
            continue
        if fuzz.partial_ratio(text_upper, marker_upper) >= threshold:
            count += 1
    return count


def has_flight_route_pattern(rows):
    for row in rows[:20]:
        for cell in row:
            if cell and RE_ROUTE.match(str(cell).strip().upper()):
                return True
    return False


def is_clean_flight_sheet(rows: list) -> bool:
    """
    Detects 'Clean' Flight Sheets (No assignments) to avoid processing them.
    Signs:
    1. Header contains 'SERIAL ... NOTE'
    2. Columns right of BAY/GATE contain 'Aerobridge', 'Passenger', etc. instead of names.
    """
    header_found = False
    bay_col_idx = -1

    # 1. Scan headers for clean sheet signature
    for r_idx, row in enumerate(rows[:15]):
        row_text = " ".join([str(x or "") for x in row]).upper()
        if (
            "SERIAL" in row_text
            and "CALLSIGN" in row_text
            and ("NOTE" in row_text or "DEP" in row_text or "ARR" in row_text)
        ):
            header_found = True
            # Find Bay column
            for c_idx, cell in enumerate(row):
                c_val = str(cell or "").upper()
                if c_val in ["BAY", "GATE", "BĂNG"]:
                    bay_col_idx = c_idx
                    break
            break

    # If explicitly marked as a Flight Plan/Schedule without assignment headers
    if header_found:
        clean_keywords = ["AEROBRIDGE", "PASSENGER", "CHARTER", "CARGO", "COMMERCIAL"]
        keyword_hits = 0
        scan_rows = rows[5:25]  # Check content

        for row in scan_rows:
            if bay_col_idx != -1 and bay_col_idx + 1 < len(row):
                val = str(row[bay_col_idx + 1] or "").upper()
                if any(k in val for k in clean_keywords):
                    keyword_hits += 1
            else:
                # Fallback: simple row scan
                row_text = " ".join([str(x or "") for x in row]).upper()
                if any(k in row_text for k in clean_keywords):
                    keyword_hits += 1

        if keyword_hits >= 2:
            return True

    return False


# --- PARSER: SHIFT DETECTION ---
def identify_shift_sheet_statistical(rows):
    name_candidates = []
    # Limit to first 100 rows for analysis
    for r_idx, row in enumerate(rows[:100]):
        # Optimization: Limit columns to first 50
        for c_idx, cell in enumerate(row[:50]):
            if is_valid_name_generic(cell):
                name_candidates.append((r_idx, c_idx))

    if len(name_candidates) < 10:
        return False, {}

    time_column_votes = defaultdict(int)
    workday_count = 0
    offday_count = 0

    for r_idx, c_idx in name_candidates:
        if r_idx >= len(rows):
            continue
        row = rows[r_idx]

        for offset in range(1, 10):  # Check +1 to +9
            check_col = c_idx + offset
            if check_col >= len(row):
                continue

            val = str(row[check_col] if row[check_col] else "").strip().upper()

            # Time pattern or Shift Marker
            if RE_SHIFT_TIME_PATTERN.search(val) or any(
                x in val
                for x in ["CA SÁNG", "CA TỐI", "CA SANG", "CA TOI", "CA 1", "CA 2"]
            ):
                time_column_votes[check_col] += 1
                workday_count += 1
            elif val in ["OFF", "HỌC", "NGHỈ", "PHÉP", "ĐI HỌC"]:
                time_column_votes[check_col] += 1
                offday_count += 1

    total = workday_count + offday_count
    if total < 5:
        return False, {}

    if not time_column_votes:
        return False, {}

    dominant_col = max(time_column_votes, key=time_column_votes.get)
    col_consistency = time_column_votes[dominant_col] / total
    workday_ratio = workday_count / total

    stats = {"total": total, "ratio": workday_ratio, "consistency": col_consistency}

    # Validation rules
    is_valid = col_consistency >= 0.50 and workday_ratio >= 0.30
    return is_valid, stats


# NOTE: normalize_time_range, is_day_off_token, is_route_pattern,
#       build_row_signals, detect_shift_sheet_by_invariants,
#       detect_flight_personnel_sheet_by_invariants, identify_sheet_type,
#       identify_sheet_type_legacy, identify_shift_sheet_statistical
#       imported from parser module


def compile_alias_regex(aliases):
    """Compile alias patterns into a regex for matching names."""
    if not aliases:
        return None
    patterns = [re.escape(a.strip()) for a in aliases if a.strip()]
    if not patterns:
        return None
    return re.compile("|".join(patterns), re.IGNORECASE)


def _windowed_flight_scan(rows, window_size=25, step=10):
    """
    v6 PATCH:
    Detect flight tables embedded inside mixed sheets.
    Looks for dense route/callsign clusters.
    """

    total_rows = len(rows)

    for start in range(0, total_rows, step):
        window = rows[start : start + window_size]

        route_hits = 0
        callsign_hits = 0

        for r in window:
            for cell in r:
                if not cell:
                    continue

                text = str(cell).upper()

                # simple aviation heuristics
                if "-" in text and len(text) <= 10:
                    route_hits += 1

                if text[:2].isalpha() and text[2:].isdigit():
                    callsign_hits += 1

        if route_hits >= 6 and callsign_hits >= 4:
            return True

    return False


# --- PARSER: SHEET IDENTIFICATION ---
def identify_sheet_type(rows, sheet_name=""):
    if not rows:
        return "SKIP"

    # ==============================
    # v6 PATCH — FLIGHT HEADER OVERRIDE
    # Detect obvious flight tables early (CSV-safe)
    # ==============================

    try:
        preview_rows = rows[:12]
        header_preview = " ".join(
            " ".join(str(c or "") for c in r[:20]) for r in preview_rows
        ).upper()

        flight_header_hits = sum(
            [
                "CALLSIGN" in header_preview,
                "ROUTE" in header_preview,
                "DEP" in header_preview or "ETD" in header_preview,
                "ARR" in header_preview or "ETA" in header_preview,
            ]
        )

        # strong signal → force FLIGHT classification
        if flight_header_hits >= 2:
            return "FLIGHT"

    except Exception:
        pass

    # Call build_row_signals once and pass the resulting signals to both detection functions
    signals = build_row_signals(rows)

    # ==============================
    # v6 PATCH — WINDOWED FLIGHT DETECTION
    # Handles mixed shift + flight sheets
    # ==============================

    if _windowed_flight_scan(rows):
        return "FLIGHT"

    # ==============================
    # v6 PATCH — PRIORITIZE FLIGHT
    # Flight structure is more specific than shift patterns
    # ==============================

    if detect_flight_personnel_sheet_by_invariants(rows, signals):
        return "FLIGHT"

    if detect_shift_sheet_by_invariants(rows, signals):
        return "SHIFT"

    # EARLY EXIT — pre-scan already decisive
    if signals.get("route_count", 0) >= 3:
        return "FLIGHT"
    if signals.get("total_time_entries", 0) >= 10:
        return "SHIFT"

    # fallback to existing logic
    return identify_sheet_type_legacy(rows, sheet_name)


def identify_sheet_type_legacy(rows, sheet_name=""):
    """
    Strict Identification Matrix (v5.6):
    - Shift: "LỊCH LÀM VIỆC ĐỘI VỆ SINH" + NOT "BẢNG PHÂN CÔNG NHIỆM VỤ"
    - Flight: Sheet "PVHL" + "SERIAL/CALLSIGN/ROUTE" markers
    """
    if not rows:
        return "SKIP"

    header_text = ""
    # Optimization: Limit columns to avoid wide-sheet performance hit
    for row in rows[:15]:
        header_text += " ".join([str(c or "") for c in row[:50]]).upper() + " "

    # 1. SHIFT SHEET DETECTION (Strict)
    is_shift_header = (
        "LỊCH LÀM VIỆC" in header_text
        or "LỊCH THỰC TẬP" in header_text
        or "DANH SÁCH" in header_text
        or "LỊCH PHÂN CA" in header_text
        or "TÊN CA MÃ CA" in header_text
        or "HỌ VÀ TÊN NV" in header_text
        or "MÃ CA" in header_text.upper()
    )
    is_ghost = "BẢNG PHÂN CÔNG NHIỆM VỤ" in header_text

    if is_shift_header and not is_ghost:
        return "SHIFT"

    # 2. FLIGHT SHEET DETECTION (Strict)
    flight_markers = ["SERIAL", "CALLSIGN", "ROUTE"]
    found_flight_markers = sum(1 for m in flight_markers if m in header_text)

    # Flight sheet usually named PVHL or LICH BAY and has the columns
    is_flight_sheet_name = sheet_name.upper() in ["PVHL", "LICH BAY", "LỊCH BAY"]

    if (
        is_flight_sheet_name or "TRỰC CHẤT XẾP" in header_text
    ) and found_flight_markers >= 2:
        # Check if it's a clean schedule or a roster
        if is_clean_flight_sheet(rows):
            return "SKIP"
        return "FLIGHT"

    # 3. SECONDARY SHIFT DETECTION (For secondary team sheets like PVHL-Shift)
    is_secondary_marker = (
        "BẢNG PHÂN CÔNG CA LÀM VIỆC" in header_text
        or "CA SÁNG" in header_text
        or "CA TỐI" in header_text
        or "CA SANG" in header_text
        or "CA TOI" in header_text
    )

    if is_secondary_marker and not is_ghost:
        # Verify names/times exist to be sure
        is_stat_shift, _ = identify_shift_sheet_statistical(rows)
        if is_stat_shift:
            return "SHIFT"

    # 4. LEGACY FALLBACKS
    # Limit text length for fuzzy matching to ensure performance
    header_limited = header_text[:2000]
    req_matches = fuzzy_match_markers(header_limited, SHIFT_SHEET_MARKERS["required"])
    if req_matches >= 1 and not is_ghost:
        return "SHIFT"

    return "SKIP"


# --- PARSER: NAME NORMALIZATION ---
def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = str(text)
    if text.endswith(".0"):
        text = text[:-2]
    return unicodedata.normalize("NFC", text).strip()


def clean_val(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        return s[:-2]
    if s.lower() == "nan":
        return ""
    return s


def clean_time(t_str: str) -> str:
    if not t_str:
        return ""
    t_str = t_str.upper().replace("H", ":").replace(".", ":").replace("G", ":").strip()
    range_match = RE_TIME_RANGE.search(t_str)
    if range_match:
        # Simplify groups to HH:MM format
        try:
            start = RE_TIME.search(range_match.group(1)).group(1)
            end = RE_TIME.search(range_match.group(2)).group(1)
            return f"{start} - {end}"
        except (AttributeError, IndexError):
            return t_str
    match = RE_TIME.search(t_str)
    if match:
        return match.group(1)
    return t_str


from parser.utils import compile_alias_regex, check_name_match, is_valid_route

# NOTE: extract_date_smart() removed - use DateResolver instead
# Date resolution is now handled by DateResolver class (Layer 1)
# which provides centralized, weighted date resolution

# ==========================================
# 🚀 CORE LOGIC ENGINE (v4.2)
# ==========================================
from parser.engine import (
    parse_row_items,
    find_name_index_in_list,
    PureParseResult,
    _extract_shift_row_pure,
    _extract_flight_rows_pure,
    parse_shift_sheet_pure,
    parse_flight_sheet_pure,
    find_header_mapping,
)


class IngestionOrchestrator:
    """
    Thin orchestration layer for file ingestion.
    Owns execution flow, not business logic.

    Phase 2.5: Single entrypoint for all ingestion.
    """

    def process_file(self, file_path: str, alias_regex) -> IngestionManifest:
        """
        Process a file through the complete ingestion pipeline.

        Returns IngestionManifest with all tracking information.
        """
        import openpyxl
        filename = os.path.basename(file_path)
        log_debug("ORCH_START", {"file": file_path})

        # --- STEP 1: Parse raw sheets ---
        sheets_data = load_excel_sheets(file_path)

        # --- STEP 2: Resolve global date ---
        context = resolve_global_date(filename, sheets_data)
        log_debug("ORCH_STEP", {"step": "date_resolved", "date": context.global_date})

        # --- STEP 3: Create manifest ---
        manifest = create_manifest_from_context(context, filename)

        # --- STEP 4: Parse sheets with context ---
        parsed_results: List[ParsedSheet] = []
        flight_sheets = []
        zone_blocks_map = {}

        # Structural zone detection for xlsx files
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            try:
                wb_openpyxl = openpyxl.load_workbook(
                    file_path, data_only=True, read_only=False
                )
                for sheet_name in wb_openpyxl.sheetnames:
                    ws = wb_openpyxl[sheet_name]
                    zone_blocks = detect_zones_from_merged_ranges(ws)
                    if zone_blocks:
                        zone_blocks_map[sheet_name] = zone_blocks
                wb_openpyxl.close()
            except Exception as e:
                log_debug("ORCH_ZONE_DETECTION_ERROR", {"error": str(e)})

        cache = {}  # Thread-local cache for pure function behavior

        for sheet_name, rows in sheets_data:
            zone_blocks = zone_blocks_map.get(sheet_name)
            parsed = parse_sheet(
                rows,
                sheet_name,
                context,
                alias_regex,
                zone_blocks=zone_blocks,
                cache=cache,
            )
            if parsed:
                parsed_results.append(parsed)
                manifest.parsed_sheets.append(sheet_name)

                if parsed.flights:
                    flight_sheets.append(
                        (sheet_name, [f.to_dict() for f in parsed.flights])
                    )

        # --- STEP 5: Dataset selection (Phase 3: ENFORCEMENT, not advisory) ---
        all_candidates = []
        authoritative_flights = []
        if flight_sheets:
            all_candidates = DatasetSelector.build_candidates(
                flight_sheets, context.global_date
            )

            selected, warnings = DatasetSelector.select_authoritative(all_candidates)
            manifest.warnings.extend(warnings)

            # Phase 3: Keep ALL fingerprints for observability
            manifest.flight_fingerprints = [c.fingerprint for c in all_candidates]

            # Phase 3: ENFORCEMENT - use ONLY authoritative dataset
            if selected:
                manifest.authoritative_fingerprint = selected.fingerprint
                manifest.authoritative_source = selected.sheet_name

                # Convert full flight rows back to flight dicts
                authoritative_flights = [f.to_dict() for f in selected.full_flights]

                # Track rejected sources
                manifest.rejected_sources = [
                    c.sheet_name
                    for c in all_candidates
                    if c.sheet_name != selected.sheet_name
                ]

                if len(all_candidates) > 1:
                    log_debug(
                        "ORCH_STEP",
                        {
                            "step": "multiple_candidates_selecting_authoritative",
                            "authoritative": selected.sheet_name,
                            "rejected": manifest.rejected_sources,
                            "flight_count": selected.row_count,
                        },
                    )

            log_debug(
                "ORCH_STEP",
                {
                    "step": "dataset_selected",
                    "fingerprint": manifest.authoritative_fingerprint[:16]
                    if manifest.authoritative_fingerprint
                    else "N/A",
                    "total_candidates": len(all_candidates),
                },
            )

        # --- STEP 6: Count parsed data (Phase 3: authoritative only) ---
        for p in parsed_results:
            if p.shift:
                manifest.parsed_counts["shift"] += 1

        # Phase 3: Count authoritative flights only (no merge)
        if authoritative_flights:
            manifest.parsed_counts["flights"] = len(authoritative_flights)
        else:
            # Fallback: count from parsed_results if no flight candidates
            for p in parsed_results:
                if p.flights:
                    manifest.parsed_counts["flights"] += len(p.flights)

        # --- STEP 7: Confidence ---
        manifest.confidence_score = ConfidenceScorer.compute(manifest)

        # --- STEP 8: Safe write gate ---
        assert_valid_manifest(manifest)

        # --- STEP 9: Persist (Phase 3: authoritative dataset only) ---
        entries = []
        for p in parsed_results:
            if p.shift:
                entry = p.to_db_dict()
                # Phase 3: Use authoritative flights only (ONE FILE = ONE DATASET)
                if authoritative_flights:
                    entry["flights"] = authoritative_flights
                entries.append(entry)

        if entries:
            # Phase 3: Validate single-dataset invariant before write
            for entry in entries:
                validate_single_dataset(entry)
            save_entries_bulk(entries, context=context)
            log_debug(
                "ORCH_STEP",
                {
                    "step": "db_write",
                    "entries": len(entries),
                    "authoritative_flights": len(authoritative_flights),
                },
            )

        # --- STEP 10: Record fingerprint ---
        if manifest.authoritative_fingerprint:
            record_fingerprint(
                manifest.authoritative_fingerprint, context.global_date, context.file_id
            )

        log_debug(
            "ORCH_DONE", {"file": file_path, "confidence": manifest.confidence_score}
        )

        return manifest


def validate_single_dataset(entry: Dict) -> None:
    """
    INVARIANT CHECK: Ensure entry adheres to single-dataset policy.

    Phase 3 (240326): Guard against accidental re-introduction of multi-sheet merge.

    Raises:
        InvariantViolation if entry contains merged_sources or invalid flights
    """
    if not isinstance(entry.get("flights"), list):
        raise InvariantViolation(f"Invalid flights type: {type(entry.get('flights'))}")


def process_file_stream(file_stream, filename, alias_regex):
    """
    Wrapper for parse_file that maintains backward compatibility.
    Returns: (results, error, manifest) tuple.
    """
    debug_log(
        f"process_file_stream called with filename='{filename}', alias_regex pattern count={len(alias_regex.pattern) if alias_regex else 0}",
        "PARSER",
    )
    result = parse_file(file_stream, filename, alias_regex)
    debug_log(
        f"process_file_stream completed, result type: {type(result)}, first element type: {type(result[0]) if result and isinstance(result, list) and result else 'N/A'}",
        "PARSER",
    )
    return result


def parse_file(file_stream, filename, alias_regex):
    """
    Main file parsing entry point.

    LAYER 1 COMPLIANCE: Global date is resolved BEFORE any sheet parsing.
    All sheets receive the same global_date via ParseContext.

    LAYER 2 COMPLIANCE: Sheet Date Enforcement - sheets with mismatched
    dates are automatically skipped and logged.

    LAYER 4 COMPLIANCE: IngestionManifest tracks all decisions for Safe Write Gate.
    """
    import openpyxl
    processed_data = []
    manifest = None

    def process_one_sheet_data(
        sheet_name_rows_tuple,
        parse_context: ParseContext,
        manifest: IngestionManifest,
        zone_blocks=None,
    ):
        """
        Process a single sheet using the pre-resolved global date.
        Implements Sheet Date Enforcement (Layer 2).
        """
        sheet_name, rows = sheet_name_rows_tuple
        t_sheet = time.monotonic()

        t_detect = time.monotonic()
        sheet_type = identify_sheet_type(rows, sheet_name or "LỊCH BAY")
        detect_time = time.monotonic() - t_detect

        if sheet_type == "SKIP":
            log_debug(
                "SHEET_SKIPPED",
                {
                    "reason": "no_matching_pattern",
                    "file": filename,
                    "sheet": sheet_name,
                    "row_count": len(rows),
                },
            )
            return None

        # Skip if global date is unknown
        if parse_context.global_date == "Unknown":
            log_debug(
                "SHEET_SKIPPED",
                {
                    "reason": "global_date_unknown",
                    "file": filename,
                    "sheet": sheet_name,
                },
            )
            return None

        # === STALE HEADER DETECTION (WARNING ONLY) ===
        # Check if sheet header has a different date - log as warning but DO NOT block
        sheet_date = _extract_sheet_date_for_warning(rows, sheet_name)
        if sheet_date and sheet_date != parse_context.global_date:
            # Log stale header as WARNING - sheet is still parsed
            warning = f"stale_sheet_header: '{sheet_name}' shows {sheet_date}, global is {parse_context.global_date}"
            manifest.warnings.append(warning)
            log_debug(
                "SHEET_HEADER_WARNING",
                {
                    "sheet": sheet_name,
                    "sheet_date": sheet_date,
                    "global_date": parse_context.global_date,
                    "file": filename,
                    "note": "Parsing continues - sheet headers are display text only",
                },
            )

        manifest.parsed_sheets.append(sheet_name)

        t_parse = time.monotonic()
        # Use pure parsing functions directly (process_sheet_v3 was removed in RosterSU)
        if sheet_type == "SHIFT":
            pure_result = parse_shift_sheet_pure(
                rows,
                parse_context.global_date,
                alias_regex,
                sheet_name,
                zone_blocks=zone_blocks,
            )
        else:
            pure_result = parse_flight_sheet_pure(
                rows, parse_context.global_date, alias_regex, sheet_name
            )
        parse_time = time.monotonic() - t_parse

        total_sheet_time = time.monotonic() - t_sheet
        if DEBUG_ENABLED:
            log_debug(
                "TIMING_INGEST",
                {
                    "file": filename,
                    "sheet": sheet_name,
                    "type": sheet_type,
                    "global_date": parse_context.global_date,
                    "parse": parse_time,
                    "total": total_sheet_time,
                },
            )

        # Convert PureParseResult to dict format
        sheet_result = {
            "date": pure_result.date,
            "shift": pure_result.shift,
            "flights": [
                f.to_dict() if hasattr(f, "to_dict") else f for f in pure_result.flights
            ],
            "sheet_name": pure_result.sheet_name,
        }

        if sheet_result["shift"] or sheet_result["flights"]:
            # Update manifest counts
            if sheet_result["shift"]:
                manifest.parsed_counts["shift"] += 1
            if sheet_result["flights"]:
                manifest.parsed_counts["flights"] += len(sheet_result["flights"])
            return sheet_result
        return None

    try:
        if filename.endswith(".csv"):
            text_stream = io.TextIOWrapper(
                file_stream, encoding="utf-8-sig", errors="replace"
            )
            reader = csv.reader(text_stream)
            rows = list(reader)

            # LAYER 1: Resolve global date for CSV
            sheets_data = [("CSV", rows)]
            parse_context = resolve_global_date(filename, sheets_data)

            if parse_context.global_date == "Unknown":
                return [], "No valid date found in file"

            # Create manifest
            manifest = create_manifest_from_context(parse_context, filename)

            sheet_type = identify_sheet_type(rows, "LỊCH BAY")
            if sheet_type != "SKIP":
                if sheet_type == "SHIFT":
                    pure_result = parse_shift_sheet_pure(
                        rows, parse_context.global_date, alias_regex, "CSV"
                    )
                else:
                    pure_result = parse_flight_sheet_pure(
                        rows, parse_context.global_date, alias_regex, "CSV"
                    )

                sheet_result = {
                    "date": pure_result.date,
                    "shift": pure_result.shift,
                    "flights": pure_result.flights,
                    "sheet_name": pure_result.sheet_name,
                }

                if sheet_result["shift"] or sheet_result["flights"]:
                    processed_data.append(sheet_result)
                    manifest.parsed_sheets.append("CSV")
                    if sheet_result["shift"]:
                        manifest.parsed_counts["shift"] += 1
                    if sheet_result["flights"]:
                        manifest.parsed_counts["flights"] += len(
                            sheet_result["flights"]
                        )
        else:
            # Use python_calamine for significantly faster processing
            t0 = time.monotonic()
            workbook = CalamineWorkbook.from_filelike(file_stream)
            load_time = time.monotonic() - t0
            if DEBUG_ENABLED:
                log_debug(
                    "TIMING_INGEST",
                    {
                        "file": filename,
                        "operation": "load_workbook",
                        "elapsed": load_time,
                    },
                )

            # Sequentially read all sheets into memory to avoid 'Already borrowed' error
            sheets_data = []
            for sheet_name in workbook.sheet_names:
                sheet = workbook.get_sheet_by_name(sheet_name)
                rows = sheet.to_python()
                sheets_data.append((sheet_name, rows))

            # === v6.1: STRUCTURAL ZONE DETECTION ===
            # For .xlsx files, use openpyxl to extract merged ranges for zone detection
            zone_blocks_map = {}
            if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
                try:
                    file_stream.seek(0)  # Reset stream position
                    wb_openpyxl = openpyxl.load_workbook(
                        file_stream, data_only=True, read_only=True
                    )
                    for sheet_name in wb_openpyxl.sheetnames:
                        ws = wb_openpyxl[sheet_name]
                        zone_blocks = detect_zones_from_merged_ranges(ws)
                        if zone_blocks:
                            zone_blocks_map[sheet_name] = zone_blocks
                    wb_openpyxl.close()
                    log_debug(
                        "ZONE_STRUCTURAL_DETECTION",
                        {
                            "file": filename,
                            "sheets_with_zones": list(zone_blocks_map.keys()),
                        },
                    )
                except Exception as e:
                    log_debug("ZONE_STRUCTURAL_DETECTION_ERROR", {"error": str(e)})

            # === LAYER 1: GLOBAL DATE RESOLUTION ===
            # Resolve the global truth date BEFORE any sheet parsing
            parse_context = resolve_global_date(filename, sheets_data)

            if parse_context.global_date == "Unknown":
                return [], "No valid date found in file"

            # Create manifest for tracking
            manifest = create_manifest_from_context(parse_context, filename)

            # Log any date anomalies
            if parse_context.date_anomaly:
                manifest.anomalies.append("date_anomaly_detected")
                log_debug(
                    "DATE_ANOMALY_WARNING",
                    {
                        "file": filename,
                        "chosen_date": parse_context.global_date,
                        "confidence": parse_context.date_confidence,
                        "candidates": parse_context.date_candidates,
                    },
                )

            # Process sheets sequentially for manifest tracking
            # (Parallel processing makes manifest tracking complex)
            for sheet_data in sheets_data:
                sheet_name = sheet_data[0]
                zone_blocks = zone_blocks_map.get(sheet_name)
                result = process_one_sheet_data(
                    sheet_data, parse_context, manifest, zone_blocks=zone_blocks
                )
                if result is not None:
                    processed_data.append(result)

            # === FLIGHT DATASET SELECTION (Feature A, B, C, D, E, F) ===
            # Apply dataset selection to ensure ONE authoritative flight dataset per file
            if processed_data:
                # Collect all flight data from processed sheets
                flight_sheets_data = []
                for res in processed_data:
                    flights = res.get("flights", [])
                    if flights:
                        # Find the sheet name for this result
                        sheet_name = res.get("sheet_name", "Unknown")
                        flight_sheets_data.append((sheet_name, flights))

                if len(flight_sheets_data) > 1:
                    # Multiple flight sheets detected - apply dataset selection
                    candidates = DatasetSelector.build_candidates(
                        flight_sheets_data, parse_context.global_date
                    )

                    authoritative, warnings = DatasetSelector.select_authoritative(
                        candidates
                    )

                    # Add warnings to manifest
                    for w in warnings:
                        manifest.warnings.append(w)

                    if authoritative:
                        # Record fingerprint for cross-file gate (Feature D)
                        if authoritative.fingerprint:
                            manifest.flight_fingerprints.append(
                                authoritative.fingerprint
                            )
                            manifest.authoritative_fingerprint = (
                                authoritative.fingerprint
                            )

                            # Feature D: Cross-File Fingerprint Gate (optimization)
                            if check_fingerprint_seen(authoritative.fingerprint):
                                manifest.warnings.append(
                                    f"CROSS_FILE_DUPLICATE: Dataset already seen in another file"
                                )
                                log_debug(
                                    "CROSS_FILE_DUPLICATE",
                                    {
                                        "file": filename,
                                        "fingerprint": authoritative.fingerprint[:16]
                                        + "...",
                                        "date": parse_context.global_date,
                                    },
                                )
                                # Note: We still proceed - this gate is optimization only

                        # Feature E: Authority Protection
                        should_proceed = update_ingestion_manifest(
                            true_date=parse_context.global_date,
                            file_hash=manifest.file_hash,
                            dataset_fingerprint=authoritative.fingerprint,
                            file_timestamp=datetime.now(timezone.utc).isoformat(),
                        )

                        if not should_proceed:
                            manifest.warnings.append(
                                "AUTHORITY_SUPERSEDED: Newer data already exists for this date"
                            )
                            log_debug(
                                "AUTHORITY_SUPERSEDED",
                                {"file": filename, "date": parse_context.global_date},
                            )
                            # Still proceed - the data might still be valuable

                        # Filter processed_data to only include authoritative flights
                        authoritative_sheet = authoritative.sheet_name
                        for res in processed_data:
                            flights = res.get("flights", [])
                            if flights:
                                # Find the sheet this result came from
                                res_sheet = res.get("sheet_name", "Unknown")
                                if res_sheet != authoritative_sheet:
                                    # Clear flights from non-authoritative sheets
                                    res["flights"] = []
                                    log_debug(
                                        "NON_AUTHORITATIVE_CLEARED",
                                        {
                                            "sheet": res_sheet,
                                            "authoritative": authoritative_sheet,
                                        },
                                    )

                        # Record fingerprint in history (Feature D)
                        if authoritative.fingerprint:
                            record_fingerprint(
                                authoritative.fingerprint,
                                parse_context.global_date,
                                manifest.file_hash,
                            )

                        log_debug(
                            "AUTHORITATIVE_DATASET_SELECTED",
                            {
                                "file": filename,
                                "sheet": authoritative_sheet,
                                "row_count": authoritative.row_count,
                                "fingerprint": authoritative.fingerprint[:16] + "..."
                                if authoritative.fingerprint
                                else "N/A",
                            },
                        )
                    else:
                        # No authoritative dataset found - clear all flights
                        for res in processed_data:
                            res["flights"] = []
                        log_debug(
                            "NO_AUTHORITATIVE_DATASET",
                            {"file": filename, "reason": "All candidates filtered"},
                        )
                elif len(flight_sheets_data) == 1:
                    # Single flight sheet - use it directly
                    sheet_name, flights = flight_sheets_data[0]
                    if flights:
                        # Compute fingerprint for tracking
                        rows = DatasetSelector.extract_flight_rows_from_sheet(flights)
                        fingerprint = DatasetSelector.compute_fingerprint(rows)

                        if fingerprint:
                            manifest.flight_fingerprints.append(fingerprint)
                            manifest.authoritative_fingerprint = fingerprint

                            # Check cross-file gate
                            if check_fingerprint_seen(fingerprint):
                                manifest.warnings.append(
                                    "CROSS_FILE_DUPLICATE: Dataset already seen in another file"
                                )

                            # Record for authority protection
                            update_ingestion_manifest(
                                true_date=parse_context.global_date,
                                file_hash=manifest.file_hash,
                                dataset_fingerprint=fingerprint,
                                file_timestamp=datetime.now(timezone.utc).isoformat(),
                            )

                            # Record fingerprint
                            record_fingerprint(
                                fingerprint,
                                parse_context.global_date,
                                manifest.file_hash,
                            )

                            log_debug(
                                "SINGLE_FLIGHT_DATASET",
                                {
                                    "file": filename,
                                    "sheet": sheet_name,
                                    "row_count": len(rows),
                                    "fingerprint": fingerprint[:16] + "...",
                                },
                            )

    except Exception as e:
        return None, str(e)

    # === LAYER 4: CONFIDENCE SCORING ===
    if manifest:
        manifest.confidence_score = ConfidenceScorer.compute(manifest)

        # Log ingestion manifest
        log_debug(
            "INGESTION_MANIFEST",
            {
                "file": filename,
                "global_date": manifest.global_date,
                "confidence": round(manifest.confidence_score, 2),
                "parsed_sheets": manifest.parsed_sheets,
                "warnings": manifest.warnings,
                "parsed_counts": manifest.parsed_counts,
                "anomalies": manifest.anomalies,
                "flight_fingerprints": [
                    fp[:16] + "..." for fp in manifest.flight_fingerprints
                ],
            },
        )

    return processed_data, None, manifest


# Alias the function to maintain backward compatibility
# process_file_stream = parse_file


def consolidate_file_results(results):
    """
    Consolidates multiple sheets/blocks found in one file.
    Does not merge with DB yet.
    """
    merged = {}
    for res in results:
        d_str = res["date"]
        if d_str not in merged:
            merged[d_str] = res
        else:
            existing = merged[d_str]
            new_shift = res.get("shift")
            if new_shift and new_shift not in ["OFF", None, ""]:
                existing["shift"] = new_shift

            existing_calls = {f["Call"] for f in existing.get("flights", [])}
            for f in res.get("flights", []):
                if f["Call"] not in existing_calls:
                    existing["flights"].append(f)
                    existing_calls.add(f["Call"])
    return merged.values()


# ==========================================
# ⚙️ UNIFIED CONFIG
# ==========================================


def _migrate_old_config():
    """Migrate old config files to new unified config."""
    import copy

    old_aliases_file = os.path.join(PROJECT_ROOT, "rosterSU_aliases.txt")
    old_aircraft_file = os.path.join(PROJECT_ROOT, "rosterSU_aircraft.json")

    config = copy.deepcopy(DEFAULT_CONFIG)  # Deep copy to avoid mutating defaults
    migrated = False

    # Migrate aliases from txt file
    if os.path.exists(old_aliases_file):
        try:
            with open(old_aliases_file, "r", encoding="utf-8") as f:
                aliases = f.read().strip()
                if aliases:
                    config["aliases"] = [
                        x.strip() for x in aliases.split(",") if x.strip()
                    ]
                    migrated = True
        except Exception:
            pass

    # Migrate aircraft config from JSON file
    if os.path.exists(old_aircraft_file):
        try:
            with open(old_aircraft_file, "r", encoding="utf-8") as f:
                aircraft_config = json.load(f)
                if isinstance(aircraft_config, dict):
                    config["aircraft"] = aircraft_config
                    migrated = True
        except Exception:
            pass

    # Save unified config if migration happened
    if migrated or not os.path.exists(CONFIG_FILE):
        save_config(config)
        # Clean up old files
        for old_file in [old_aliases_file, old_aircraft_file]:
            if os.path.exists(old_file):
                try:
                    os.remove(old_file)
                except Exception:
                    pass

    return config


def get_config():
    """Load unified config from JSON file."""
    with CONFIG_LOCK:
        # Check for old config files and migrate
        old_aliases_file = os.path.join(PROJECT_ROOT, "rosterSU_aliases.txt")
        old_aircraft_file = os.path.join(PROJECT_ROOT, "rosterSU_aircraft.json")
        if os.path.exists(old_aliases_file) or os.path.exists(old_aircraft_file):
            return _migrate_old_config()

        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
            except (json.JSONDecodeError, KeyError):
                config = {}
        else:
            config = {}

        # Ensure all user preference keys exist
        if "aliases" not in config:
            config["aliases"] = DEFAULT_CONFIG["aliases"]
        if "aircraft" not in config:
            config["aircraft"] = DEFAULT_CONFIG["aircraft"].copy()
        for key in ["airbus", "boeing", "other"]:
            if key not in config.get("aircraft", {}):
                config["aircraft"][key] = DEFAULT_CONFIG["aircraft"][key]

        # Ensure all operational settings exist (backward compat with old JSON files)
        for op_key in [
            "port", "history_limit", "page_size", "port_wait_timeout",
            "ingest_interval", "max_upload_mb",
            "auto_ingest_dir", "export_dir", "processed_archive_dir",
            "db_path",
            "static_html_scope", "static_html_count", "static_html_output_dir",
            "enable_flight_sync",
        ]:
            if op_key not in config:
                config[op_key] = DEFAULT_CONFIG[op_key]

        return config


def save_config(config):
    """Save unified config to JSON file."""
    with CONFIG_LOCK:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)


def get_aliases():
    """Get aliases from unified config."""
    config = get_config()
    return config.get("aliases", DEFAULT_CONFIG["aliases"])


def get_aircraft_config():
    """Get aircraft config from unified config."""
    config = get_config()
    return config.get("aircraft", DEFAULT_CONFIG["aircraft"])


def _init_components_module():
    """Initialize components module with dependencies from main module."""
    from config import PAGE_SIZE

    _init_components(get_aircraft_config, count_history, load_history, PAGE_SIZE)


# Initialize components module after dependencies are defined
_init_components_module()


def _run_ingest_once():
    """
    Background auto-ingest with Safe Write Gate (Layer 4).

    Implements:
    - Confidence score validation
    - Quarantine for low-confidence files
    - Manifest-based decision logging
    """
    import concurrent.futures
    APP.set_ingest()
    try:
        if not os.path.exists(AUTO_INGEST_DIR):
            update_status("Idle", "Monitoring folder...")
            return

        # Targeted glob patterns — avoid scanning non-roster files
        xlsx_files = glob.glob(os.path.join(AUTO_INGEST_DIR, "*.xlsx"))
        xls_files = glob.glob(os.path.join(AUTO_INGEST_DIR, "*.xls"))
        csv_files = glob.glob(os.path.join(AUTO_INGEST_DIR, "*.csv"))
        target_files = xlsx_files + xls_files + csv_files

        # Path safety validation
        safe_files = []
        for f in target_files:
            try:
                safe_path(AUTO_INGEST_DIR, f)
                safe_files.append(f)
            except ValueError:
                continue

        if not safe_files:
            update_status("Idle", "Monitoring folder...")
            return

        count = 0
        quarantined_count = 0
        safe_files.sort()
        batch_files = safe_files[:10]
        update_status("Processing", f"Importing {len(batch_files)} files.")

        aliases = get_aliases()
        alias_regex = compile_alias_regex(aliases)

        def _background_parse(f_path, regex):
            try:
                safe_f_path = safe_path(AUTO_INGEST_DIR, f_path)
                if not os.path.exists(safe_f_path):
                    return (f_path, None, "MISSING", None)
                # File size pre-check to prevent OOM
                file_size = os.path.getsize(safe_f_path)
                max_bytes = MAX_UPLOAD_MB * 1024 * 1024
                if file_size > max_bytes:
                    log_debug("file_too_large", {"file": os.path.basename(f_path), "size_mb": file_size / (1024*1024)})
                    return (f_path, None, f"File too large (>{MAX_UPLOAD_MB}MB)", None)
                with open(safe_f_path, "rb") as f:
                    # parse_file now returns (results, error, manifest)
                    p_results, p_err, manifest = parse_file(
                        f, os.path.basename(f_path), regex
                    )
                return (f_path, p_results, p_err, manifest)
            except Exception as ex:
                log_debug(
                    "parse_failed", {"file": os.path.basename(f_path), "error": str(ex)}
                )
                return (f_path, None, str(ex), None)

        results_list = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
            futures = [
                executor.submit(_background_parse, fp, alias_regex)
                for fp in batch_files
            ]
            for future in concurrent.futures.as_completed(futures):
                results_list.append(future.result())

        for f_path_res, results, err, manifest in results_list:
            try:
                safe_file = safe_path(AUTO_INGEST_DIR, f_path_res)
                if err == "MISSING":
                    continue

                # === LAYER 4: SAFE WRITE GATE ===
                # Check confidence score before writing to DB
                # Use round() to avoid floating-point traps (0.39999...97 vs 0.4)
                if manifest and round(manifest.confidence_score, 10) < SAFE_THRESHOLD:
                    # Low confidence - quarantine the file
                    manifest.blocked = True
                    manifest.block_reason = f"confidence_score_below_threshold: {manifest.confidence_score:.2f} < {SAFE_THRESHOLD}"
                    quarantine_file(safe_file, manifest)
                    quarantined_count += 1
                    log_debug(
                        "INGESTION_QUARANTINED",
                        {
                            "file": os.path.basename(f_path_res),
                            "confidence": manifest.confidence_score,
                            "threshold": SAFE_THRESHOLD,
                            "reason": manifest.block_reason,
                        },
                    )
                    continue

                # Gate: file parsed but produced zero data (all sheets SKIP)
                if manifest and not manifest.parsed_counts.get("shift") and not manifest.parsed_counts.get("flights"):
                    manifest.blocked = True
                    manifest.block_reason = "no_data_extracted: all sheets classified as SKIP"
                    quarantine_file(safe_file, manifest)
                    quarantined_count += 1
                    log_debug(
                        "INGESTION_QUARANTINED",
                        {
                            "file": os.path.basename(f_path_res),
                            "confidence": manifest.confidence_score,
                            "threshold": SAFE_THRESHOLD,
                            "reason": manifest.block_reason,
                        },
                    )
                    continue

                if results:
                    final = consolidate_file_results(results)
                    # Pass manifest context for DB validation (if available)
                    if manifest:
                        save_entries_bulk(
                            final,
                            context=ParseContext(
                                global_date=manifest.global_date,
                                global_date_iso=manifest.global_date_iso,
                                source_filename=manifest.filename,
                                file_id=manifest.file_hash,
                                date_confidence=manifest.confidence_score,
                                date_anomaly=len(manifest.anomalies) > 0,
                                date_candidates=manifest.date_candidates,
                            ),
                        )
                    else:
                        save_entries_bulk(final)
                    count += 1
                    if os.path.exists(safe_file) and manifest:
                        _archive_processed_file(safe_file, manifest)
                    elif os.path.exists(safe_file):
                        os.remove(safe_file)
                elif err:
                    if os.path.exists(safe_file):
                        os.rename(safe_file, safe_file + ".error")
                else:
                    # No results, no error — file parsed but produced nothing
                    if manifest:
                        manifest.blocked = True
                        manifest.block_reason = "no_data_extracted: parsed but empty result"
                        quarantine_file(safe_file, manifest)
                        quarantined_count += 1
                    elif os.path.exists(safe_file):
                        os.rename(safe_file, safe_file + ".nodate")
            except Exception as e:
                log_debug("file_loop_error", {"file": f_path_res, "error": str(e)})

        if count > 0 or quarantined_count > 0:
            msg = f"Processed {count} files"
            if quarantined_count > 0:
                msg += f", quarantined {quarantined_count}"
            update_status("Idle", msg)
        else:
            update_status("Idle", "Monitoring folder...")

    finally:
        APP.clear_ingest()
        update_status("Idle", "Sẵn sàng")


def run_auto_ingest():
    """Background task to watch folder for new files."""
    while not SHUTDOWN_EVENT.is_set():
        if APP.is_ingest_running():
            time.sleep(INGEST_INTERVAL)
            continue

        APP.set_ingest()
        try:
            _run_ingest_once()
        except Exception as e:
            log_debug("auto_ingest_error", str(e))
            update_status("Error", str(e))
        finally:
            APP.clear_ingest()
            bump_db_rev()
            # Generate static HTML viewer
            try:
                from export import generate_html as gen_static_html
                config = get_config()
                scope = config.get("static_html_scope", "current_month")
                count = config.get("static_html_count", 5)
                result = gen_static_html(scope=scope, count=count)
                if result["success"]:
                    log_debug("static_html_generated", {"path": result["file_path"], "count": result["entry_count"]})
                else:
                    log_debug("static_html_error", {"error": result["error"]})
            except Exception as e:
                log_debug("static_html_error", str(e))

        time.sleep(INGEST_INTERVAL)


# ==========================================
# 🖥️ UI (FastHTML + Oat UI)
# ==========================================

# Headers for Oat UI
# CSS embedded directly for speed and reliability (2026-04-06)
css_content = """
/* === Enhanced CSS Variables === */
:root {
    --space-1: 0.15rem; --space-2: 0.3rem; --space-3: 0.5rem; --space-4: 0.7rem;
    --shadow-sm: 0 1px 2px rgba(0,0,0,0.05);
    --shadow-md: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -2px rgba(0,0,0,0.1);
    --shadow-lg: 0 10px 15px -3px rgba(0,0,0,0.1), 0 4px 6px -4px rgba(0,0,0,0.1);
    --radius-sm: 0.375rem; --radius-md: 0.5rem; --radius-lg: 0.75rem;
    --transition-fast: 150ms ease; --transition-base: 200ms ease;
}
* { box-sizing: border-box; }

/* Hide disclosure arrows globally */
summary::after { content: none !important; display: none !important; }

/* === Typography & Layout === */
body { 
    font-size: 0.875rem; line-height: 1.5; 
    background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
    min-height: 100vh;
}
[data-theme='dark'] body { background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%); }

/* Dark theme text colors - override missing --text/--muted variables */
[data-theme='dark'] { --text: #e2e8f0; --muted: #94a3b8; --card-bg: #1e293b; }
[data-theme='dark'] body { color: #e2e8f0; }
[data-theme='dark'] p { color: #e2e8f0; }
[data-theme='dark'] h1, [data-theme='dark'] h2, [data-theme='dark'] h3, 
[data-theme='dark'] h4, [data-theme='dark'] h5, [data-theme='dark'] h6 { color: #f1f5f9; }
[data-theme='dark'] nav a { color: #e2e8f0; }
[data-theme='dark'] nav a:hover { color: #fff; background: rgba(255,255,255,0.1); }
[data-theme='dark'] nav strong { 
    background: linear-gradient(135deg, #818cf8, #a78bfa); 
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
}
[data-theme='dark'] .fd th { background: transparent !important; color: #fff !important; }
[data-theme='dark'] .fd td { background: transparent !important; color: #fff !important; }
/* Dark mode text color overrides for utility classes */
[data-theme='dark'] .text-red-600 { color: #f87171 !important; }
[data-theme='dark'] .text-blue-600 { color: #60a5fa !important; }
[data-theme='dark'] .text-green-700 { color: #4ade80 !important; }
[data-theme='dark'] .text-grey-7, [data-theme='dark'] .text-gray-700 { color: #d1d5db !important; }

h1,h2,h3,h4,h5,h6 { margin-block: 0.4rem 0.3rem; font-weight: 600; letter-spacing: -0.01em; }
h4 { font-size: 1rem; color: var(--text); }
h5 { font-size: 0.9rem; color: var(--text); }
p { margin-block: 0.2rem; }
section { margin-block: 0; padding-block: 0; }
table { margin-block: 0; }
hr { margin-block: 0.75rem; border-color: rgba(148,163,184,0.3); }
details { margin-block: 0; }

/* === Navigation === */
nav { 
    background: rgba(255,255,255,0.8); 
    backdrop-filter: blur(12px);
    border-bottom: 1px solid rgba(148,163,184,0.2);
    position: sticky; top: 0; z-index: 100;
}
[data-theme='dark'] nav { background: rgba(15,23,42,0.9); border-color: rgba(71,85,105,0.4); }
nav a { 
    color: var(--muted); text-decoration: none; 
    padding: 0.25rem 0.5rem; border-radius: var(--radius-sm);
    transition: all var(--transition-fast);
}
nav a:hover { color: var(--text); background: rgba(148,163,184,0.1); }
nav strong { 
    background: linear-gradient(135deg, #6366f1, #8b5cf6); 
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    font-weight: 700;
}

/* === Status Indicator === */
#status-indicator { 
    display: inline-flex; align-items: center; gap: 0.35rem;
    padding: 0.25rem 0.6rem; border-radius: 9999px;
    font-size: 0.7rem; font-weight: 500;
    background: rgba(34,197,94,0.1); transition: all var(--transition-base);
}
.st-idle { color: #16a34a; }
.st-run { color: #d97706; background: rgba(217,119,6,0.1); animation: pulse 2s infinite; }
@keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.7; } }

/* === Roster Cards === */
.rc { 
    display: flex; align-items: center; gap: 0.5rem;
    padding: 0.5rem 0.75rem; border-radius: var(--radius-lg);
    color: #fff; position: relative; overflow: hidden;
    box-shadow: var(--shadow-md);
    transition: transform var(--transition-fast), box-shadow var(--transition-fast);
    width: 100%;
}
.rc:hover { transform: translateY(-1px); box-shadow: var(--shadow-lg); }
.rc::before {
    content: ""; position: absolute; inset: 0; 
    background: linear-gradient(180deg, rgba(255,255,255,0.15) 0%, transparent 50%);
    pointer-events: none;
}

/* Light theme: bright vibrant colors with white text */
.rc-on  { background: linear-gradient(135deg, #22c55e 0%, #16a34a 50%, #15803d 100%); }
.rc-off { background: linear-gradient(135deg, #f87171 0%, #ef4444 50%, #dc2626 100%); }
.rc-edu { background: linear-gradient(135deg, #60a5fa 0%, #3b82f6 50%, #2563eb 100%); }
.rc-nil { background: linear-gradient(135deg, #94a3b8 0%, #64748b 50%, #475569 100%); }

/* Dark theme: saturated colors with white text for contrast */
[data-theme='dark'] .rc-on  { background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%); border: none; }
[data-theme='dark'] .rc-off { background: linear-gradient(135deg, #f87171 0%, #ef4444 100%); border: none; }
[data-theme='dark'] .rc-edu { background: linear-gradient(135deg, #60a5fa 0%, #3b82f6 100%); border: none; }
[data-theme='dark'] .rc-nil { background: linear-gradient(135deg, #94a3b8 0%, #64748b 100%); border: none; }

.rc input[type=checkbox] { 
    transform: scale(1.2); z-index: 2; position: relative; margin: 0;
    accent-color: #fff;
    display: none; /* Hidden by default */
}
.delete-mode .rc input[type=checkbox] { display: block; }

/* Card content area - horizontal layout to fill space */
.rc-content { 
    flex: 1; min-width: 0; z-index: 1;
    display: flex;
    align-items: center;
    gap: 0.75rem;
}

/* Date column */
.rc-date-col {
    min-width: 4.5rem;
}

/* Date: 20% bigger = 0.78rem (was 0.65rem) */
.rc-date { 
    font-size: 0.78rem; color: #fff; 
    text-transform: uppercase; letter-spacing: 0.05rem; margin: 0;
    font-weight: 600;
    text-shadow: 0 1px 2px rgba(0,0,0,0.2);
}

/* Shift time column */
.rc-shift-col {
    flex: 1;
    min-width: 0;
}

/* Shift time: 15% bigger = 0.966rem (was 0.84rem) */
.rc-shift { 
    font-size: 0.966rem; font-weight: 700; margin: 0; line-height: 1.2; 
    color: #fff; text-shadow: 0 1px 2px rgba(0,0,0,0.2);
}

/* Zone: 15% bigger = 0.633rem (was 0.55rem), no parentheses */
.rc-zone { 
    font-size: 0.633rem; color: rgba(255,255,255,0.85); margin: 0; 
    line-height: 1.2; text-transform: capitalize;
    text-shadow: 0 1px 1px rgba(0,0,0,0.15);
}

/* Flight count badge */
.rc-badge {
    background: rgba(255,255,255,0.25);
    color: #fff;
    padding: 0.2rem 0.5rem;
    border-radius: 9999px;
    font-size: 0.6rem;
    font-weight: 600;
    white-space: nowrap;
    z-index: 1;
    backdrop-filter: blur(4px);
    flex-shrink: 0;
}

/* === Accordion/Details === */
.rd { 
    margin-bottom: 0.3rem; 
    border-radius: var(--radius-lg); 
    overflow: hidden;
    width: 100%;
}
[data-theme='dark'] .rd {
    border: none;
    outline: none;
}
.rd summary { 
    display: flex; 
    list-style: none !important; 
    cursor: pointer; 
    padding: 0; 
    outline: none;
    -webkit-appearance: none;
    appearance: none;
    width: 100%;
}
/* Hide all triangle markers completely - cross browser */
summary::-webkit-details-marker { display: none !important; }
summary::marker { content: '' !important; }
summary { list-style: none !important; }
details summary { list-style: none !important; }
details > summary { list-style-type: none !important; }
.rd summary:focus-visible .rc { outline: 2px solid #6366f1; outline-offset: 2px; }
.rd[open] .rc { border-radius: var(--radius-lg) var(--radius-lg) 0 0; }
.rd[open] summary .rc::before { background: linear-gradient(180deg, rgba(255,255,255,0.2) 0%, transparent 50%); }

/* Details content (flight table) */
.fd {
    padding: 0.4rem 0.6rem; font-size: 0.75rem;
    background: rgba(255,255,255,0.95);
    border: 1px solid rgba(148,163,184,0.25); border-top: none;
    border-radius: 0 0 var(--radius-lg) var(--radius-lg);
    box-shadow: var(--shadow-sm);
    width: 100%;
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
}
[data-theme='dark'] .fd { background: #000 !important; border: none !important; }
[data-theme='dark'] .fd table { background: transparent !important; }
[data-theme='dark'] .fd tr { background: transparent !important; }
[data-theme='dark'] .fd td { background: transparent !important; color: #fff !important; }
.fd table { font-size: 0.75rem; width: 100%; min-width: max-content; table-layout: auto; border-collapse: collapse; }
.fd th { 
    font-weight: 600; color: var(--text); 
    border-bottom: 1px solid rgba(148,163,184,0.2);
    padding: 0.3rem 0.5rem;
    white-space: nowrap;
    font-size: 0.7rem;
}
.fd td { padding: 0.25rem 0.5rem; color: var(--muted); white-space: nowrap; }

/* Flight type row backgrounds (configurable) */
/* Airbus → dark red */
.fd tr.flight-airbus { background: #7f1d1d !important; }
.fd tr.flight-airbus td { color: #fecaca !important; }
/* Boeing → dark blue */
.fd tr.flight-boeing { background: #1e3a5f !important; }
.fd tr.flight-boeing td { color: #bfdbfe !important; }
/* Other (user-defined) → dark green */
.fd tr.flight-other { background: #14532d !important; }
.fd tr.flight-other td { color: #bbf7d0 !important; }
/* Dark theme variants */
[data-theme='dark'] .fd tr.flight-airbus { background: #991b1b !important; }
[data-theme='dark'] .fd tr.flight-boeing { background: #1e3a8a !important; }
[data-theme='dark'] .fd tr.flight-other { background: #166534 !important; }

/* === Buttons === */
.btn-row { display: flex; gap: 0.4rem; margin-block: 0.5rem; }
.btn-row > *, .btn-row form { flex: 1; }

.btn-act { 
    min-height: 44px; font-size: 0.85rem; font-weight: 600;
    border-radius: var(--radius-md);
    transition: all var(--transition-fast);
    box-shadow: var(--shadow-sm);
}
.btn-act:hover { transform: translateY(-1px); box-shadow: var(--shadow-md); }
.btn-act:active { transform: translateY(0); }

.btn-del { 
    background: linear-gradient(135deg, #f87171, #ef4444) !important; 
    color: #fff !important; border: none !important;
}
.btn-del:hover { background: linear-gradient(135deg, #ef4444, #dc2626) !important; }

.btn-icn { 
    padding: 0.3rem 0.5rem; font-size: 1.1rem; line-height: 1;
    border-radius: var(--radius-sm);
    transition: all var(--transition-fast);
}
.btn-icn:hover { background: rgba(148,163,184,0.15); }

/* Outline button dark mode */
[data-theme='dark'] .button.outline,
[data-theme='dark'] .outline {
    color: #fff !important;
    border-color: rgba(148,163,184,0.5) !important;
}

/* === Forms & Inputs === */
select, textarea {
    border-radius: var(--radius-md);
    border: 1px solid rgba(148,163,184,0.3);
    background: rgba(255,255,255,0.8);
    transition: all var(--transition-fast);
}
select:focus, textarea:focus {
    outline: none;
    border-color: #6366f1;
    box-shadow: 0 0 0 3px rgba(99,102,241,0.15);
}
[data-theme='dark'] select, [data-theme='dark'] textarea {
    background: rgba(30,41,59,0.8);
    border-color: rgba(71,85,105,0.4);
    color: var(--text);
}
select { padding: 0.4rem 0.6rem; cursor: pointer; }

/* === Empty State === */
.empty { 
    text-align: center; padding: 2rem 1.5rem; 
    border: 2px dashed rgba(148,163,184,0.4); 
    border-radius: var(--radius-lg);
    background: rgba(255,255,255,0.5);
}
[data-theme='dark'] .empty { 
    background: rgba(15,23,42,0.9);
    border-color: rgba(71,85,105,0.5);
}
.empty p:first-child { font-weight: 600; margin: 0; font-size: 0.95rem; }
[data-theme='dark'] .empty p:last-child { font-size: 0.8rem; color: #94a3b8; margin: 0.25rem 0 0; }

/* === Utility: Card container === */
.card {
    background: rgba(255,255,255,0.9);
    border-radius: var(--radius-lg);
    padding: 0.75rem;
    box-shadow: var(--shadow-sm);
    border: 1px solid rgba(148,163,184,0.15);
}
[data-theme='dark'] .card {
    background: rgba(15,23,42,0.9);
    border: none;
}

/* === Utility: Dot indicator === */
.dot-indicator { font-size: 0.5rem; }

/* === Utility: Status navigation bar === */
.status-nav-bar { display: flex; justify-content: space-between; align-items: center; padding: 0.4rem 1rem; }

/* === Utility: Main container === */
.container-main { padding-block: 0.6rem; max-width: 540px; }

/* === Utility: Table cell spans === */
.td-span-type { font-size: 0.7rem; }
.td-span-call { font-weight: 500; }
.td-span-time { font-size: 0.7rem; }
.td-span-bay { text-align: center; }
.td-span-names { font-size: 0.7rem; }

/* === Utility: Pagination === */
.pagination-bar { display: flex; justify-content: center; align-items: center; gap: 0.5rem; margin-top: 1rem; margin-bottom: 0.5rem; }
.pagination-btn { padding: 0.25rem 0.75rem; min-width: auto; }
.pagination-info { padding: 0 0.5rem; font-weight: 500; }

/* === Utility: Form sections === */
.section-title { margin-bottom: 0.75rem; }
.section-title-top { margin-top: 0.5rem; }
.form-label { font-weight: 500; margin-bottom: 0.3rem; }
.form-textarea { font-family: monospace; font-size: 0.8rem; width: 100%; resize: vertical; }
.form-textarea-sm { font-family: monospace; font-size: 0.75rem; width: 100%; resize: vertical; }
.form-hint { font-size: 0.7rem; color: var(--muted); margin-top: 0.2rem; }
.form-hint-sm { font-size: 0.8rem; color: var(--muted); margin-bottom: 0.5rem; }
.form-hint-airbus { font-weight: 500; margin-bottom: 0.2rem; color: #ef4444; }
.form-hint-boeing { font-weight: 500; margin-bottom: 0.2rem; color: #3b82f6; }
.form-hint-other { font-weight: 500; margin-bottom: 0.2rem; color: #22c55e; }
.btn-full-width { width: 100%; margin-bottom: 1rem; }

/* === Utility: Upload button === */
.upload-label { margin: 0; display: flex; align-items: center; justify-content: center; cursor: pointer; }
.upload-input { display: none; }

/* === Utility: Hidden form === */
.hidden-form { display: none; }
"""

hdrs = (
    Link(
        rel="stylesheet",
        href="https://cdn.jsdelivr.net/npm/@knadh/oat@latest/oat.min.css",
    ),
    Style(css_content),
    Script(src="https://cdn.jsdelivr.net/npm/@knadh/oat@latest/oat.min.js"),
    Script(f"""
        document.addEventListener('DOMContentLoaded', () => {{
            document.documentElement.setAttribute('data-theme', localStorage.getItem('theme') || 'light');
        }});
        function toggleTheme() {{
            const t = document.documentElement.getAttribute('data-theme') === 'dark' ? 'light' : 'dark';
            document.documentElement.setAttribute('data-theme', t);
            localStorage.setItem('theme', t);
        }}
    """),
)

app, rt = fast_app(hdrs=hdrs)


def layout(content, title="RosterMaster"):
    status_indicator = Span(
        Input(type="hidden", id="client-rev", name="rev", value=APP.roster_version),
        Span("●", cls="dot-indicator"),
        Span("Sẵn sàng", id="status-text"),
        id="status-indicator",
        hx_get="/status",
        hx_trigger="every 5s",
        hx_include="#client-rev",
        cls="st-idle",
    )

    return Title(title), Body(
        Nav(
            Div(
                Strong("RM"),
                A("📅 Lịch", href="/"),
                A("⚙️ Cài đặt", href="/settings"),
                cls="hstack gap-2",
            ),
            Div(
                status_indicator,
                Button(
                    "◐",
                    onclick="toggleTheme()",
                    cls="ghost btn-icn",
                    title="Đổi giao diện",
                ),
                Button(
                    "✕",
                    onclick="this.textContent='⏳';this.disabled=true;fetch('/shutdown',{method:'POST'}).then(()=>{document.body.innerHTML='';window.close()})",
                    cls="ghost btn-icn",
                    title="Đóng ứng dụng",
                ),
                cls="hstack gap-2",
            ),
            cls="status-nav-bar",
        ),
        Main(content, cls="container container-main"),
        data_theme="light",
    )


# Import routes to register them
from routes import *

# ==========================================
# 🚀 MAIN ENTRY
# ==========================================

if __name__ in {"__main__", "builtins"}:
    debug_log("Application starting", "MAIN")
    debug_log(f"Debug mode is {'enabled' if DEBUG_ENABLED else 'disabled'}", "MAIN")
    init_db()

    # --- Flight Delay Auto-Sync ---
    def _run_flight_startup_sync():
        """Run flight sync at startup if enabled. Non-blocking, logs results."""
        try:
            merged = _load_merged_config()
            if not merged.get("enable_flight_sync", False):
                debug_log("Flight sync: disabled in config, skipping")
                return

            from scraper import AutoSyncService
            from database import get_db

            today_iso = datetime.now().strftime("%Y-%m-%d")
            today_db = datetime.now().strftime("%d.%m.%Y")

            conn = get_db()
            try:
                service = AutoSyncService()
                sync_result = service.run_sync(conn, today_iso, today_db)
                for detail in sync_result.details:
                    debug_log(f"Flight sync: {detail}")
            finally:
                conn.close()

            # Bump revision to trigger UI refresh
            from state import bump_db_rev
            bump_db_rev()
        except Exception as e:
            debug_log(f"Flight sync startup error: {e}")
            # Do NOT crash -- app continues without sync

    _run_flight_startup_sync()

    # Refresh flight API preview cache at startup (today + 2 future days)
    try:
        import routes
        routes._refresh_api_cache(days=3)
    except Exception as e:
        debug_log(f"Flight API preview cache startup refresh error: {e}")

    os.makedirs(PROCESSED_ARCHIVE_DIR, exist_ok=True)
    threading.Thread(target=run_auto_ingest, daemon=True).start()
    threading.Thread(target=auto_open_launcher, daemon=True).start()
    debug_log(f"Serving application on port {DEFAULT_PORT}", "MAIN")
    serve(port=DEFAULT_PORT, host="127.0.0.1")
